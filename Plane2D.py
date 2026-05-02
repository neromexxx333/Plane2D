import base64
import io
import tempfile
import textwrap
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.tri as mtri
import numpy as np
import pandas as pd
import streamlit as st
from matplotlib.animation import FuncAnimation, PillowWriter
from matplotlib.cm import ScalarMappable
from matplotlib.colors import LinearSegmentedColormap, Normalize
from matplotlib.offsetbox import AnchoredOffsetbox, HPacker, TextArea, VPacker
from matplotlib.patches import Rectangle
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


GAUSS_POINTS_2X2 = [
    (-1.0 / np.sqrt(3.0), -1.0 / np.sqrt(3.0), 1.0),
    (1.0 / np.sqrt(3.0), -1.0 / np.sqrt(3.0), 1.0),
    (1.0 / np.sqrt(3.0), 1.0 / np.sqrt(3.0), 1.0),
    (-1.0 / np.sqrt(3.0), 1.0 / np.sqrt(3.0), 1.0),
]
LINE_GAUSS_POINTS_2 = [
    (-1.0 / np.sqrt(3.0), 1.0),
    (1.0 / np.sqrt(3.0), 1.0),
]
EDGE_NODE_MAP = {
    1: (0, 1),
    2: (1, 2),
    3: (2, 3),
    4: (3, 0),
}
EDGE_NATURAL_MAP = {
    1: lambda s: (s, -1.0),
    2: lambda s: (1.0, s),
    3: lambda s: (s, 1.0),
    4: lambda s: (-1.0, s),
}
TEMPLATE_FILENAME = "template_plane_stress_quad4.xlsx"
EXAMPLE_FILENAME = "input_plane_stress_quad.xlsx"
RESULTS_FILENAME = "hasil_plane2d_analisis.xlsx"
KGCM2_TO_MPA = 0.1
MAX_EXPORT_MATRIX_CELLS = 120000
STRESS_RESULT_COLUMNS = [
    "sx",
    "sy",
    "txy",
    "s1",
    "s2",
    "s3",
    "mean_stress",
    "hydrostatic_mean_3d",
    "tau_max",
    "tau_min",
    "von_mises",
    "I1",
    "sqrt_J2",
    "k_dp",
    "dp_value",
    "yield_function",
]
STRESS_SQUARED_RESULT_COLUMNS = ["J2"]
STRESS_CONTOUR_OPTIONS = {
    "Von Mises": "von_mises",
    "sx": "sx",
    "sy": "sy",
    "txy": "txy",
    "s1": "s1",
    "s2": "s2",
    "Tegangan rata-rata": "mean_stress",
}
STRESS_CONTOUR_LABELS = {
    column_name: label for label, column_name in STRESS_CONTOUR_OPTIONS.items()
}
STRESS_CONTOUR_CMAP = LinearSegmentedColormap.from_list(
    "stress_yellow_red",
    ["#ffe45e", "#ff9f1c", "#d62828"],
)
DISPLACEMENT_CONTOUR_OPTIONS = {
    "Ux": "Ux",
    "Uy": "Uy",
}
DISPLACEMENT_CONTOUR_CMAP = LinearSegmentedColormap.from_list(
    "displacement_blue_to_red",
    ["#08306b", "#2171b5", "#6baed6", "#c6dbef", "#d62828"],
)
ANIMATION_FIGURE_SIZE = (10.5, 6.8)


def require_columns(df, required_columns, sheet_name):
    missing = [column for column in required_columns if column not in df.columns]
    if missing:
        raise ValueError(
            f"Sheet '{sheet_name}' belum memiliki kolom wajib: {', '.join(missing)}"
        )


def read_required_sheet(xls, sheet_name, required_columns):
    try:
        df = pd.read_excel(xls, sheet_name=sheet_name)
    except ValueError as exc:
        raise ValueError(f"Sheet wajib '{sheet_name}' tidak ditemukan") from exc
    require_columns(df, required_columns, sheet_name)
    return df


def read_optional_sheet(xls, sheet_name, columns):
    try:
        df = pd.read_excel(xls, sheet_name=sheet_name)
    except ValueError:
        return pd.DataFrame(columns=columns)
    for column in columns:
        if column not in df.columns:
            df[column] = 0.0
    return df


def signed_polygon_area(coords):
    x = coords[:, 0]
    y = coords[:, 1]
    return 0.5 * np.sum(x * np.roll(y, -1) - y * np.roll(x, -1))


def constitutive_matrix(E, nu, mode):
    if mode == "Plane Stress":
        factor = E / (1.0 - nu**2)
        return factor * np.array(
            [[1.0, nu, 0.0], [nu, 1.0, 0.0], [0.0, 0.0, (1.0 - nu) / 2.0]]
        )

    factor = E / ((1.0 + nu) * (1.0 - 2.0 * nu))
    return factor * np.array(
        [
            [1.0 - nu, nu, 0.0],
            [nu, 1.0 - nu, 0.0],
            [0.0, 0.0, (1.0 - 2.0 * nu) / 2.0],
        ]
    )


def shape_functions_quad4(xi, eta):
    return 0.25 * np.array(
        [
            (1.0 - xi) * (1.0 - eta),
            (1.0 + xi) * (1.0 - eta),
            (1.0 + xi) * (1.0 + eta),
            (1.0 - xi) * (1.0 + eta),
        ]
    )


def shape_derivatives_quad4(xi, eta):
    dN_dxi = 0.25 * np.array(
        [-(1.0 - eta), (1.0 - eta), (1.0 + eta), -(1.0 + eta)]
    )
    dN_deta = 0.25 * np.array(
        [-(1.0 - xi), -(1.0 + xi), (1.0 + xi), (1.0 - xi)]
    )
    return np.vstack((dN_dxi, dN_deta))


def quad4_B_matrix(coords, xi, eta):
    dN_natural = shape_derivatives_quad4(xi, eta)
    jacobian = dN_natural @ coords
    det_jacobian = np.linalg.det(jacobian)
    if det_jacobian <= 0.0:
        raise ValueError(
            "Jacobian elemen tidak valid. Pastikan urutan node elemen berlawanan arah jarum jam."
        )

    dN_global = np.linalg.inv(jacobian) @ dN_natural
    B = np.zeros((3, 8))
    for i in range(4):
        B[0, 2 * i] = dN_global[0, i]
        B[1, 2 * i + 1] = dN_global[1, i]
        B[2, 2 * i] = dN_global[1, i]
        B[2, 2 * i + 1] = dN_global[0, i]
    return B, det_jacobian


def map_natural_to_global(coords, xi, eta):
    return shape_functions_quad4(xi, eta) @ coords


def quad4_element_matrices(coords, D, thickness):
    Ke = np.zeros((8, 8))
    gauss_data = []
    B_total = np.zeros((3, 8))

    for gauss_index, (xi, eta, weight) in enumerate(GAUSS_POINTS_2X2, start=1):
        B, det_jacobian = quad4_B_matrix(coords, xi, eta)
        BTDB = B.T @ D @ B
        Ke += BTDB * det_jacobian * thickness * weight
        gauss_data.append(
            {
                "gauss_point": gauss_index,
                "xi": xi,
                "eta": eta,
                "weight": weight,
                "det_jacobian": det_jacobian,
                "B": B,
                "BTDB": BTDB,
            }
        )
        B_total += B

    return {
        "ke": Ke,
        "gauss_data": gauss_data,
        "B_total": B_total,
    }


def quad4_edge_load(coords, edge, qx, qy):
    if edge not in EDGE_NODE_MAP:
        raise ValueError("Nomor edge elemen harus 1, 2, 3, atau 4")

    start_node, end_node = EDGE_NODE_MAP[edge]
    jacobian_edge = np.linalg.norm(coords[end_node] - coords[start_node]) / 2.0
    traction = np.array([qx, qy], dtype=float)
    fe = np.zeros(8)

    for s, weight in LINE_GAUSS_POINTS_2:
        xi, eta = EDGE_NATURAL_MAP[edge](s)
        N = shape_functions_quad4(xi, eta)
        N_matrix = np.zeros((2, 8))
        for i in range(4):
            N_matrix[0, 2 * i] = N[i]
            N_matrix[1, 2 * i + 1] = N[i]
        fe += N_matrix.T @ traction * jacobian_edge * weight

    return fe


def element_dof_map(node_ids, node_index):
    dof_map = []
    for node_id in node_ids:
        idx = node_index[node_id]
        dof_map.extend([2 * idx, 2 * idx + 1])
    return dof_map


def element_dof_labels(node_ids):
    labels = []
    for node_id in node_ids:
        labels.extend([f"u{node_id}", f"v{node_id}"])
    return labels


def global_dof_labels(node_ids):
    return element_dof_labels(node_ids)


def b_total_row_labels():
    return ["ex", "ey", "gxy"]


def matrix_to_dataframe(matrix, row_labels, col_labels):
    return pd.DataFrame(matrix, index=row_labels, columns=col_labels)


def vector_to_dataframe(vector, labels, column_name):
    return pd.DataFrame({"DOF": labels, column_name: vector})


def show_dataframe(df, *, keep_left_column=False, **kwargs):
    if keep_left_column:
        st.dataframe(df, **kwargs)
        return
    st.dataframe(df, hide_index=True, **kwargs)


def assemble_element_global_matrix(ke, dof_map, total_dof):
    kg_element = np.zeros((total_dof, total_dof))
    kg_element[np.ix_(dof_map, dof_map)] = ke
    return kg_element


def normalize_angle_deg(angle_deg):
    return ((angle_deg + 90.0) % 180.0) - 90.0


def principal_angle_display_deg(angle_deg):
    return float(angle_deg)


def principal_plot_direction(angle_deg):
    angle_value = float(angle_deg)
    if np.isclose(angle_value, 0.0):
        return "Sejajar sumbu +x"
    if angle_value > 0.0:
        return "Berlawanan arah jarum jam"
    return "Searah jarum jam"


def perpendicular_angle_from_principal_deg(angle_deg):
    return float(angle_deg) + 90.0


def add_counterclockwise_principal_angle_columns(df):
    display_df = df.copy()
    if "principal_angle_1_deg" in display_df.columns:
        display_df["principal_angle_1_ccw_deg"] = display_df[
            "principal_angle_1_deg"
        ].apply(principal_angle_display_deg)
        display_df["principal_angle_plot_direction"] = display_df[
            "principal_angle_1_deg"
        ].apply(principal_plot_direction)
        display_df["principal_angle_perpendicular_deg"] = display_df[
            "principal_angle_1_deg"
        ].apply(perpendicular_angle_from_principal_deg)
    if "principal_angle_2_deg" in display_df.columns:
        display_df["principal_angle_2_ccw_deg"] = display_df[
            "principal_angle_2_deg"
        ].apply(principal_angle_display_deg)
    return display_df


def convert_stress_dataframe_to_mpa(df):
    converted_df = df.copy()
    for column in STRESS_RESULT_COLUMNS:
        if column in converted_df.columns:
            converted_df[column] = converted_df[column] * KGCM2_TO_MPA
    for column in STRESS_SQUARED_RESULT_COLUMNS:
        if column in converted_df.columns:
            converted_df[column] = converted_df[column] * (KGCM2_TO_MPA**2)
    return converted_df


def convert_stress_matrix_to_mpa(matrix):
    return np.array(matrix, dtype=float) * KGCM2_TO_MPA


def transform_stress_2d(sx, sy, txy, theta_deg):
    theta = np.radians(theta_deg)
    mean_inplane = 0.5 * (sx + sy)
    diff = 0.5 * (sx - sy)
    sigma_theta = mean_inplane + diff * np.cos(2.0 * theta) + txy * np.sin(2.0 * theta)
    tau_theta = -diff * np.sin(2.0 * theta) + txy * np.cos(2.0 * theta)
    return sigma_theta, tau_theta


def workbook_principal_angle_deg(sx, sy, txy):
    angle_raw = 0.5 * np.degrees(np.arctan2(2.0 * txy, sx - sy))
    angle_option_1 = normalize_angle_deg(angle_raw)
    angle_option_2 = normalize_angle_deg(angle_raw + 90.0)
    if abs(angle_option_1) <= abs(angle_option_2):
        return angle_option_1
    return angle_option_2


def drucker_prager_parameters(fc, ft):
    denominator = np.sqrt(3.0) * (fc + ft)
    if denominator <= 0.0:
        raise ValueError("Parameter beton untuk Drucker-Prager harus lebih besar dari nol.")
    alpha = (fc - ft) / denominator
    k_value = (2.0 * fc * ft) / denominator
    return alpha, k_value


def stress_metrics(sx, sy, txy, fc, ft):
    mean_inplane = 0.5 * (sx + sy)
    radius = np.sqrt(((sx - sy) / 2.0) ** 2 + txy**2)
    sigma_1 = mean_inplane + radius
    sigma_2 = mean_inplane - radius
    sigma_3 = 0.0

    theta_p1 = workbook_principal_angle_deg(sx, sy, txy)
    theta_p2 = theta_p1 + 90.0

    tau_max = radius
    theta_tau_max = normalize_angle_deg(theta_p1 + 45.0)
    tau_min = -radius
    theta_tau_min = normalize_angle_deg(theta_p1 - 45.0)

    mean_stress = 0.5 * (sx + sy)
    hydrostatic_mean_3d = (sx + sy + sigma_3) / 3.0
    von_mises = np.sqrt(sx**2 - sx * sy + sy**2 + 3.0 * txy**2)

    I1 = sx + sy + sigma_3
    J2 = ((sx - sy) ** 2 + (sy - sigma_3) ** 2 + (sigma_3 - sx) ** 2) / 6.0 + txy**2
    sqrt_J2 = np.sqrt(max(J2, 0.0))
    alpha_dp, k_dp = drucker_prager_parameters(fc, ft)
    dp_value = sqrt_J2 + alpha_dp * I1
    yield_function = dp_value - k_dp

    return {
        "s1": sigma_1,
        "s2": sigma_2,
        "s3": sigma_3,
        "principal_angle_1_deg": normalize_angle_deg(theta_p1),
        "principal_angle_2_deg": theta_p2,
        "mean_stress": mean_stress,
        "hydrostatic_mean_3d": hydrostatic_mean_3d,
        "tau_max": tau_max,
        "theta_tau_max_deg": theta_tau_max,
        "tau_min": tau_min,
        "theta_tau_min_deg": theta_tau_min,
        "von_mises": von_mises,
        "I1": I1,
        "J2": J2,
        "sqrt_J2": sqrt_J2,
        "alpha_dp": alpha_dp,
        "k_dp": k_dp,
        "dp_value": dp_value,
        "yield_function": yield_function,
        "yield_state": "Yield" if yield_function >= 0.0 else "Elastic",
    }


def row_value(row, column_name, default_value):
    if column_name in row and pd.notna(row[column_name]):
        return float(row[column_name])
    return float(default_value)


def solve_system(K, F, prescribed_dofs):
    dof = K.shape[0]
    fixed_dofs = sorted(prescribed_dofs)
    free_dofs = [i for i in range(dof) if i not in prescribed_dofs]

    if not free_dofs:
        raise ValueError("Semua derajat bebas terkunci.")

    U = np.zeros(dof)
    if fixed_dofs:
        U[fixed_dofs] = np.array([prescribed_dofs[dof_id] for dof_id in fixed_dofs])

    Kff = K[np.ix_(free_dofs, free_dofs)]
    Ff = F[free_dofs]
    if fixed_dofs:
        Kfc = K[np.ix_(free_dofs, fixed_dofs)]
        Uc = U[fixed_dofs]
        Ff = Ff - Kfc @ Uc

    try:
        U[free_dofs] = np.linalg.solve(Kff, Ff)
    except np.linalg.LinAlgError as exc:
        raise ValueError(
            "Matriks kekakuan singular. Cek boundary condition atau konektivitas mesh."
        ) from exc

    reactions = K @ U - F
    return U, reactions


def create_results(
    nodes,
    elements,
    node_index,
    coordinates,
    mode,
    default_E,
    default_nu,
    default_t,
    concrete_fc,
    concrete_ft,
    U,
):
    displacement_rows = []
    for _, node in nodes.iterrows():
        node_id = int(node["id"])
        idx = node_index[node_id]
        ux = U[2 * idx]
        uy = U[2 * idx + 1]
        displacement_rows.append(
            {
                "node": node_id,
                "x": coordinates[idx, 0],
                "y": coordinates[idx, 1],
                "Ux": ux,
                "Uy": uy,
                "U_mag": np.hypot(ux, uy),
            }
        )

    gauss_rows = []
    element_rows = []
    for _, element in elements.iterrows():
        element_id = int(element["id"])
        node_ids = [int(element[f"n{i}"]) for i in range(1, 5)]
        coords = np.array([coordinates[node_index[node_id]] for node_id in node_ids])
        centroid = coords.mean(axis=0)
        dof_map = element_dof_map(node_ids, node_index)
        element_u = U[dof_map]

        E = row_value(element, "E", default_E)
        nu = row_value(element, "nu", default_nu)
        thickness = row_value(element, "t", default_t)
        D = constitutive_matrix(E, nu, mode)

        B_total = np.zeros((3, 8))
        for gauss_index, (xi, eta, _) in enumerate(GAUSS_POINTS_2X2, start=1):
            B, _ = quad4_B_matrix(coords, xi, eta)
            B_total += B
            strain = B @ element_u
            stress = D @ strain
            sx, sy, txy = stress
            metrics = stress_metrics(sx, sy, txy, concrete_fc, concrete_ft)
            gp_coordinates = map_natural_to_global(coords, xi, eta)
            row = {
                "element": element_id,
                "gauss_point": gauss_index,
                "xi": xi,
                "eta": eta,
                "x_gp": gp_coordinates[0],
                "y_gp": gp_coordinates[1],
                "thickness": thickness,
                "ex": strain[0],
                "ey": strain[1],
                "gxy": strain[2],
                "sx": sx,
                "sy": sy,
                "txy": txy,
                **metrics,
            }
            gauss_rows.append(row)

        element_strain = B_total @ element_u
        element_stress = D @ element_strain
        sx, sy, txy = element_stress
        element_metrics = stress_metrics(sx, sy, txy, concrete_fc, concrete_ft)

        element_rows.append(
            {
                "element": element_id,
                "x_centroid": centroid[0],
                "y_centroid": centroid[1],
                "thickness": thickness,
                "ex": element_strain[0],
                "ey": element_strain[1],
                "gxy": element_strain[2],
                "sx": sx,
                "sy": sy,
                "txy": txy,
                **element_metrics,
            }
        )

    return (
        pd.DataFrame(displacement_rows),
        pd.DataFrame(gauss_rows),
        pd.DataFrame(element_rows),
    )


def auto_scale_factor(nodes_xy, displacements):
    max_dimension = max(
        np.ptp(nodes_xy[:, 0]) if len(nodes_xy) else 0.0,
        np.ptp(nodes_xy[:, 1]) if len(nodes_xy) else 0.0,
    )
    max_displacement = np.max(np.abs(displacements)) if len(displacements) else 0.0
    if max_dimension <= 0.0 or max_displacement <= 1e-12:
        return 1.0
    return max(1.0, 0.15 * max_dimension / max_displacement)


def select_extreme_row(df, column_name, *, mode="max", absolute=False):
    if df.empty or column_name not in df.columns:
        return None

    numeric_series = pd.to_numeric(df[column_name], errors="coerce")
    valid_values = numeric_series.dropna()
    if valid_values.empty:
        return None

    if absolute:
        selected_index = valid_values.abs().idxmax()
    elif mode == "min":
        selected_index = valid_values.idxmin()
    else:
        selected_index = valid_values.idxmax()
    return df.loc[selected_index]


def build_support_reaction_detail_dataframe(reactions, prescribed_dofs, node_ids):
    rows = []
    for dof_id in sorted(prescribed_dofs):
        node_id = int(node_ids[dof_id // 2])
        component = "Rx" if dof_id % 2 == 0 else "Ry"
        rows.append(
            {
                "node": node_id,
                "component": component,
                "prescribed_displacement": prescribed_dofs[dof_id],
                "reaction": reactions[dof_id],
                "reaction_abs": abs(reactions[dof_id]),
            }
        )
    return pd.DataFrame(rows)


def build_equilibrium_dataframe(F, reactions, prescribed_dofs, relative_tolerance=1e-6):
    support_reaction_x = sum(
        reactions[dof_id] for dof_id in prescribed_dofs if dof_id % 2 == 0
    )
    support_reaction_y = sum(
        reactions[dof_id] for dof_id in prescribed_dofs if dof_id % 2 == 1
    )
    total_load_x = float(F[0::2].sum())
    total_load_y = float(F[1::2].sum())
    residual_x = total_load_x + float(support_reaction_x)
    residual_y = total_load_y + float(support_reaction_y)

    reference_x = max(abs(total_load_x), abs(support_reaction_x), 1.0)
    reference_y = max(abs(total_load_y), abs(support_reaction_y), 1.0)
    residual_ratio_x = abs(residual_x) / reference_x
    residual_ratio_y = abs(residual_y) / reference_y

    residual_norm = float(np.hypot(residual_x, residual_y))
    reference_norm = max(
        float(np.hypot(total_load_x, total_load_y)),
        float(np.hypot(support_reaction_x, support_reaction_y)),
        1.0,
    )
    residual_ratio_norm = residual_norm / reference_norm

    equilibrium_df = pd.DataFrame(
        [
            {
                "Arah": "Fx",
                "Total beban luar": total_load_x,
                "Total reaksi tumpuan": support_reaction_x,
                "Residual (beban + reaksi)": residual_x,
                "Residual relatif": residual_ratio_x,
                "Status": "OK"
                if residual_ratio_x <= relative_tolerance
                else "Warning",
            },
            {
                "Arah": "Fy",
                "Total beban luar": total_load_y,
                "Total reaksi tumpuan": support_reaction_y,
                "Residual (beban + reaksi)": residual_y,
                "Residual relatif": residual_ratio_y,
                "Status": "OK"
                if residual_ratio_y <= relative_tolerance
                else "Warning",
            },
            {
                "Arah": "Resultan",
                "Total beban luar": float(np.hypot(total_load_x, total_load_y)),
                "Total reaksi tumpuan": float(
                    np.hypot(support_reaction_x, support_reaction_y)
                ),
                "Residual (beban + reaksi)": residual_norm,
                "Residual relatif": residual_ratio_norm,
                "Status": "OK"
                if residual_ratio_norm <= relative_tolerance
                else "Warning",
            },
        ]
    )
    return equilibrium_df, bool((equilibrium_df["Status"] == "OK").all())


def build_model_summary_dataframe(
    analysis_mode,
    nodes,
    elements,
    loads,
    bc,
    edge_loads,
    default_E,
    default_nu,
    default_t,
    concrete_fc,
    concrete_ft,
    dof_total,
    prescribed_dofs,
):
    fixed_dofs = len(prescribed_dofs)
    free_dofs = max(dof_total - fixed_dofs, 0)
    return pd.DataFrame(
        [
            {"Parameter": "Tipe analisis", "Nilai": analysis_mode, "Satuan": "-"},
            {"Parameter": "Jumlah node", "Nilai": len(nodes), "Satuan": "node"},
            {"Parameter": "Jumlah elemen", "Nilai": len(elements), "Satuan": "elemen"},
            {"Parameter": "DOF total", "Nilai": dof_total, "Satuan": "DOF"},
            {"Parameter": "DOF terkunci", "Nilai": fixed_dofs, "Satuan": "DOF"},
            {"Parameter": "DOF bebas", "Nilai": free_dofs, "Satuan": "DOF"},
            {"Parameter": "Baris BC", "Nilai": len(bc), "Satuan": "baris"},
            {
                "Parameter": "Jumlah beban nodal",
                "Nilai": len(loads),
                "Satuan": "baris",
            },
            {
                "Parameter": "Jumlah beban tepi",
                "Nilai": len(edge_loads),
                "Satuan": "baris",
            },
            {
                "Parameter": "E default",
                "Nilai": float(default_E),
                "Satuan": "kg/cm2",
            },
            {"Parameter": "nu default", "Nilai": float(default_nu), "Satuan": "-"},
            {
                "Parameter": "Tebal default",
                "Nilai": float(default_t),
                "Satuan": "cm",
            },
            {
                "Parameter": "fc' beton",
                "Nilai": float(concrete_fc),
                "Satuan": "kg/cm2",
            },
            {
                "Parameter": "ft beton",
                "Nilai": float(concrete_ft),
                "Satuan": "kg/cm2",
            },
        ]
    )


def format_node_location(row):
    return (
        f"Node {int(row['node'])} "
        f"(x={float(row['x']):.3f}, y={float(row['y']):.3f})"
    )


def format_element_location(row):
    return (
        f"Elemen {int(row['element'])} "
        f"(x={float(row['x_centroid']):.3f}, y={float(row['y_centroid']):.3f})"
    )


def format_gauss_location(row):
    return (
        f"Elemen {int(row['element'])} / GP {int(row['gauss_point'])} "
        f"(x={float(row['x_gp']):.3f}, y={float(row['y_gp']):.3f})"
    )


def build_executive_summary_dataframe(
    displacement_df,
    element_results,
    gauss_results,
    support_reaction_detail_df,
):
    summary_rows = []

    max_displacement_row = select_extreme_row(displacement_df, "U_mag")
    if max_displacement_row is not None:
        summary_rows.append(
            {
                "Indikator": "Displacement maksimum",
                "Lokasi": format_node_location(max_displacement_row),
                "Nilai": float(max_displacement_row["U_mag"]),
                "Satuan": "satuan model",
                "Keterangan": "Magnitude perpindahan total terbesar.",
            }
        )

    max_ux_row = select_extreme_row(displacement_df, "Ux", absolute=True)
    if max_ux_row is not None:
        summary_rows.append(
            {
                "Indikator": "|Ux| maksimum",
                "Lokasi": format_node_location(max_ux_row),
                "Nilai": float(max_ux_row["Ux"]),
                "Satuan": "satuan model",
                "Keterangan": "Komponen perpindahan arah X terbesar secara absolut.",
            }
        )

    max_uy_row = select_extreme_row(displacement_df, "Uy", absolute=True)
    if max_uy_row is not None:
        summary_rows.append(
            {
                "Indikator": "|Uy| maksimum",
                "Lokasi": format_node_location(max_uy_row),
                "Nilai": float(max_uy_row["Uy"]),
                "Satuan": "satuan model",
                "Keterangan": "Komponen perpindahan arah Y terbesar secara absolut.",
            }
        )

    max_support_reaction_row = select_extreme_row(
        support_reaction_detail_df, "reaction", absolute=True
    )
    if max_support_reaction_row is not None:
        summary_rows.append(
            {
                "Indikator": "Reaksi tumpuan maksimum",
                "Lokasi": (
                    f"Node {int(max_support_reaction_row['node'])} / "
                    f"{max_support_reaction_row['component']}"
                ),
                "Nilai": float(max_support_reaction_row["reaction"]),
                "Satuan": "satuan gaya model",
                "Keterangan": "Reaksi per DOF terkunci terbesar secara absolut.",
            }
        )

    max_vm_element_row = select_extreme_row(element_results, "von_mises")
    if max_vm_element_row is not None:
        summary_rows.append(
            {
                "Indikator": "Von Mises maksimum elemen",
                "Lokasi": format_element_location(max_vm_element_row),
                "Nilai": float(max_vm_element_row["von_mises"]),
                "Satuan": "MPa",
                "Keterangan": "Berdasarkan hasil elemen memakai `B_total`.",
            }
        )

    max_vm_gauss_row = select_extreme_row(gauss_results, "von_mises")
    if max_vm_gauss_row is not None:
        summary_rows.append(
            {
                "Indikator": "Von Mises maksimum titik Gauss",
                "Lokasi": format_gauss_location(max_vm_gauss_row),
                "Nilai": float(max_vm_gauss_row["von_mises"]),
                "Satuan": "MPa",
                "Keterangan": "Lebih konservatif karena dievaluasi per titik integrasi.",
            }
        )

    max_tension_row = select_extreme_row(gauss_results, "s1")
    if max_tension_row is not None:
        summary_rows.append(
            {
                "Indikator": "Tegangan utama tarik maksimum",
                "Lokasi": format_gauss_location(max_tension_row),
                "Nilai": float(max_tension_row["s1"]),
                "Satuan": "MPa",
                "Keterangan": "Mengacu pada principal stress `s1`.",
            }
        )

    max_compression_row = select_extreme_row(gauss_results, "s2", mode="min")
    if max_compression_row is not None:
        summary_rows.append(
            {
                "Indikator": "Tegangan utama tekan maksimum",
                "Lokasi": format_gauss_location(max_compression_row),
                "Nilai": float(max_compression_row["s2"]),
                "Satuan": "MPa",
                "Keterangan": "Nilai paling negatif menunjukkan tekan terbesar.",
            }
        )

    max_yield_row = select_extreme_row(gauss_results, "yield_function")
    if max_yield_row is not None:
        summary_rows.append(
            {
                "Indikator": "Yield function maksimum",
                "Lokasi": format_gauss_location(max_yield_row),
                "Nilai": float(max_yield_row["yield_function"]),
                "Satuan": "MPa",
                "Keterangan": f"Status titik ini: {max_yield_row['yield_state']}.",
            }
        )

    yield_count = (
        int((gauss_results["yield_state"] == "Yield").sum())
        if "yield_state" in gauss_results.columns
        else 0
    )
    summary_rows.append(
        {
            "Indikator": "Jumlah titik Gauss status Yield",
            "Lokasi": "Seluruh model",
            "Nilai": yield_count,
            "Satuan": "titik",
            "Keterangan": (
                "Perlu review lebih lanjut."
                if yield_count > 0
                else "Semua titik Gauss masih berstatus Elastic."
            ),
        }
    )

    return pd.DataFrame(summary_rows)


def build_nodal_hotspots_dataframe(displacement_df, top_n=10):
    if displacement_df.empty:
        return displacement_df.copy()

    hotspot_df = displacement_df.copy()
    hotspot_df["Abs_Ux"] = hotspot_df["Ux"].abs()
    hotspot_df["Abs_Uy"] = hotspot_df["Uy"].abs()
    hotspot_df = hotspot_df.sort_values(
        ["U_mag", "Abs_Uy", "Abs_Ux"], ascending=[False, False, False]
    ).head(top_n)
    return hotspot_df[["node", "x", "y", "Ux", "Uy", "U_mag"]]


def build_support_hotspots_dataframe(constrained_reaction_df, displacement_df, top_n=10):
    if constrained_reaction_df.empty:
        return constrained_reaction_df.copy()

    hotspot_df = constrained_reaction_df.merge(
        displacement_df[["node", "x", "y"]],
        on="node",
        how="left",
    )
    hotspot_df["R_mag"] = np.hypot(hotspot_df["Rx"], hotspot_df["Ry"])
    hotspot_df = hotspot_df.sort_values("R_mag", ascending=False).head(top_n)
    return hotspot_df[["node", "x", "y", "Rx", "Ry", "R_mag"]]


def build_element_hotspot_dataframe(element_results, primary_column, top_n=10):
    if element_results.empty or primary_column not in element_results.columns:
        return pd.DataFrame()

    sort_columns = [primary_column]
    ascending_flags = [False]
    if primary_column != "von_mises" and "von_mises" in element_results.columns:
        sort_columns.append("von_mises")
        ascending_flags.append(False)

    hotspot_df = element_results.sort_values(
        sort_columns, ascending=ascending_flags
    ).head(top_n)
    return hotspot_df[
        [
            "element",
            "x_centroid",
            "y_centroid",
            "sx",
            "sy",
            "txy",
            "s1",
            "s2",
            "von_mises",
            "yield_function",
            "yield_state",
        ]
    ]


def build_gauss_hotspot_dataframe(gauss_results, primary_column, top_n=10):
    if gauss_results.empty or primary_column not in gauss_results.columns:
        return pd.DataFrame()

    sort_columns = [primary_column]
    ascending_flags = [False]
    if primary_column != "von_mises" and "von_mises" in gauss_results.columns:
        sort_columns.append("von_mises")
        ascending_flags.append(False)

    hotspot_df = gauss_results.sort_values(
        sort_columns, ascending=ascending_flags
    ).head(top_n)
    return hotspot_df[
        [
            "element",
            "gauss_point",
            "x_gp",
            "y_gp",
            "sx",
            "sy",
            "txy",
            "s1",
            "s2",
            "von_mises",
            "yield_function",
            "yield_state",
        ]
    ]


def style_export_worksheet(worksheet, df, *, include_index):
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    sample_size = min(len(df), 50) if hasattr(df, "__len__") else 0
    column_offset = 1
    if include_index:
        sample_index = df.index[:sample_size]
        index_width = max(
            [len(str(value)) for value in sample_index] + [12],
        )
        worksheet.column_dimensions["A"].width = min(index_width + 2, 26)
        column_offset = 2

    for column_position, column_name in enumerate(df.columns, start=column_offset):
        sample_values = df[column_name].head(sample_size).tolist()
        max_length = max(
            [len(str(column_name))] + [len(str(value)) for value in sample_values] + [10]
        )
        worksheet.column_dimensions[get_column_letter(column_position)].width = min(
            max_length + 2, 28
        )


def write_export_sheet(writer, sheet_name, df, *, include_index=False):
    df.to_excel(writer, sheet_name=sheet_name, index=include_index)
    worksheet = writer.sheets[sheet_name]
    style_export_worksheet(worksheet, df, include_index=include_index)


def build_results_workbook(sheet_specs):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_spec in sheet_specs:
            write_export_sheet(
                writer,
                sheet_spec["sheet_name"],
                sheet_spec["dataframe"],
                include_index=sheet_spec.get("include_index", False),
            )
    output.seek(0)
    return output.getvalue()


def build_plot_coordinates(coordinates, U=None, scale_factor=0.0):
    plot_coordinates = np.array(coordinates, dtype=float, copy=True)
    if U is None:
        return plot_coordinates
    displacement_matrix = np.asarray(U, dtype=float).reshape(-1, 2)
    return plot_coordinates + scale_factor * displacement_matrix


def element_node_indices_list(elements, node_index):
    return [
        [node_index[int(element[f"n{i}"])] for i in range(1, 5)]
        for _, element in elements.iterrows()
    ]


def average_element_result_to_nodes(elements, node_index, element_results, result_column):
    value_by_element = element_results.set_index("element")[result_column].to_dict()
    nodal_sum = np.zeros(len(node_index), dtype=float)
    nodal_count = np.zeros(len(node_index), dtype=float)

    for _, element in elements.iterrows():
        element_id = int(element["id"])
        if element_id not in value_by_element or pd.isna(value_by_element[element_id]):
            continue

        element_value = float(value_by_element[element_id])
        node_ids = [int(element[f"n{i}"]) for i in range(1, 5)]
        for node_id in node_ids:
            idx = node_index[node_id]
            nodal_sum[idx] += element_value
            nodal_count[idx] += 1.0

    return np.divide(
        nodal_sum,
        nodal_count,
        out=np.zeros_like(nodal_sum),
        where=nodal_count > 0.0,
    )


def build_quad_triangulation(elements, node_index, plot_coordinates):
    triangles = []
    for node_indices in element_node_indices_list(elements, node_index):
        triangles.append([node_indices[0], node_indices[1], node_indices[2]])
        triangles.append([node_indices[0], node_indices[2], node_indices[3]])

    if not triangles:
        return None

    return mtri.Triangulation(
        plot_coordinates[:, 0], plot_coordinates[:, 1], np.array(triangles, dtype=int)
    )


def draw_contour_mesh_overlay(ax, elements, node_index, plot_coordinates, bc_nodes):
    for node_indices in element_node_indices_list(elements, node_index):
        coords = np.array([plot_coordinates[idx] for idx in node_indices])
        coords_closed = np.vstack((coords, coords[0]))
        ax.plot(
            coords_closed[:, 0],
            coords_closed[:, 1],
            color="0.2",
            linewidth=0.8,
            alpha=0.55,
        )

    if bc_nodes:
        bc_coordinates = np.array(
            [plot_coordinates[node_index[node_id]] for node_id in sorted(bc_nodes)]
        )
        ax.scatter(
            bc_coordinates[:, 0],
            bc_coordinates[:, 1],
            s=28,
            facecolors="none",
            edgecolors="black",
            linewidths=0.9,
            zorder=3,
        )


def compute_animation_axis_limits(coordinates, U, scale_factor):
    frames = [
        np.asarray(coordinates, dtype=float),
        build_plot_coordinates(coordinates, U, scale_factor),
    ]
    stacked = np.vstack(frames)
    x_min = float(stacked[:, 0].min())
    x_max = float(stacked[:, 0].max())
    y_min = float(stacked[:, 1].min())
    y_max = float(stacked[:, 1].max())

    span_x = max(x_max - x_min, 1.0)
    span_y = max(y_max - y_min, 1.0)
    pad_x = 0.08 * span_x
    pad_y = 0.08 * span_y
    return (
        x_min - pad_x,
        x_max + pad_x,
        y_min - pad_y,
        y_max + pad_y,
    )


def animation_amplitude_factor(frame_index, frame_count):
    if frame_count <= 1:
        return 1.0
    return float(np.sin(np.pi * frame_index / (frame_count - 1)))


def scalar_animation_limits(base_values):
    finite_values = np.asarray(base_values, dtype=float)
    finite_values = finite_values[np.isfinite(finite_values)]
    if finite_values.size == 0:
        return -1.0, 1.0

    value_min = min(0.0, float(finite_values.min()))
    value_max = max(0.0, float(finite_values.max()))
    if np.isclose(value_min, value_max):
        delta = max(abs(value_min) * 1e-6, 1e-6)
        return value_min - delta, value_max + delta
    return value_min, value_max


def save_animation_to_gif(fig, animation, interval_ms):
    fps = max(1, int(round(1000.0 / max(interval_ms, 1))))
    with tempfile.NamedTemporaryFile(
        dir=Path.cwd(), suffix=".gif", delete=False
    ) as tmp_file:
        temp_path = Path(tmp_file.name)

    try:
        animation.save(str(temp_path), writer=PillowWriter(fps=fps), dpi=72)
        return temp_path.read_bytes()
    finally:
        if temp_path.exists():
            temp_path.unlink()
        plt.close(fig)


def render_gif(gif_bytes):
    if not gif_bytes:
        return

    encoded_gif = base64.b64encode(gif_bytes).decode("ascii")
    st.markdown(
        (
            "<div style='width:100%; text-align:center;'>"
            f"<img src='data:image/gif;base64,{encoded_gif}' "
            "style='max-width:100%; height:auto; border-radius:4px;' />"
            "</div>"
        ),
        unsafe_allow_html=True,
    )


def create_deformation_animation_gif(
    elements,
    node_index,
    coordinates,
    U,
    scale_factor,
    bc_nodes,
    frames,
    interval_ms,
):
    if len(coordinates) == 0:
        return None

    fig, ax = plt.subplots(figsize=ANIMATION_FIGURE_SIZE)
    x_min, x_max, y_min, y_max = compute_animation_axis_limits(
        coordinates, U, scale_factor
    )
    ax.set_xlim(x_min, x_max)
    ax.set_ylim(y_min, y_max)
    ax.set_aspect("equal")
    ax.set_xlabel("X")
    ax.set_ylabel("Y")
    ax.grid(True, alpha=0.2)

    element_indices = element_node_indices_list(elements, node_index)
    original_coordinates = np.asarray(coordinates, dtype=float)
    for node_indices in element_indices:
        coords = np.array([original_coordinates[idx] for idx in node_indices])
        coords_closed = np.vstack((coords, coords[0]))
        ax.plot(
            coords_closed[:, 0],
            coords_closed[:, 1],
            color="0.6",
            linestyle="--",
            linewidth=1.0,
        )

    deformed_lines = []
    for _ in element_indices:
        (line,) = ax.plot([], [], color="#005bbb", linewidth=2.0)
        deformed_lines.append(line)

    bc_indices = [node_index[node_id] for node_id in sorted(bc_nodes)]
    bc_scatter = None
    if bc_indices:
        initial_bc_coords = original_coordinates[bc_indices]
        bc_scatter = ax.scatter(
            initial_bc_coords[:, 0],
            initial_bc_coords[:, 1],
            s=32,
            c="#b22222",
            marker="s",
            zorder=3,
        )

    title = ax.set_title("Animasi Deformasi")

    def update(frame_index):
        frame_scale = scale_factor * animation_amplitude_factor(frame_index, frames)
        animated_coordinates = build_plot_coordinates(coordinates, U, frame_scale)

        for line, node_indices in zip(deformed_lines, element_indices):
            coords = np.array([animated_coordinates[idx] for idx in node_indices])
            coords_closed = np.vstack((coords, coords[0]))
            line.set_data(coords_closed[:, 0], coords_closed[:, 1])

        if bc_scatter is not None:
            bc_scatter.set_offsets(animated_coordinates[bc_indices])

        title.set_text(f"Animasi Deformasi - Skala saat ini: {frame_scale:.3g}")
        artists = list(deformed_lines)
        if bc_scatter is not None:
            artists.append(bc_scatter)
        artists.append(title)
        return artists

    animation = FuncAnimation(
        fig,
        update,
        frames=max(int(frames), 2),
        interval=int(interval_ms),
        blit=False,
        repeat=True,
    )
    return save_animation_to_gif(fig, animation, interval_ms)


def create_scalar_contour_animation_gif(
    elements,
    node_index,
    coordinates,
    U,
    scale_factor,
    bc_nodes,
    base_values,
    result_label,
    unit_label,
    cmap,
    frames,
    interval_ms,
    title_prefix,
):
    if len(coordinates) == 0:
        return None

    value_min, value_max = scalar_animation_limits(base_values)
    x_min, x_max, y_min, y_max = compute_animation_axis_limits(
        coordinates, U, scale_factor
    )
    fig = plt.figure(figsize=ANIMATION_FIGURE_SIZE)
    grid_spec = fig.add_gridspec(1, 2, width_ratios=[24, 1.1], wspace=0.08)
    ax = fig.add_subplot(grid_spec[0, 0])
    cax = fig.add_subplot(grid_spec[0, 1])
    norm = Normalize(vmin=value_min, vmax=value_max)
    colorbar = fig.colorbar(ScalarMappable(norm=norm, cmap=cmap), cax=cax)
    colorbar.set_label(f"{result_label} ({unit_label})")

    def update(frame_index):
        ax.clear()

        amplitude = animation_amplitude_factor(frame_index, frames)
        animated_coordinates = build_plot_coordinates(
            coordinates, U, scale_factor * amplitude
        )
        animated_values = np.asarray(base_values, dtype=float) * amplitude
        triangulation = build_quad_triangulation(
            elements, node_index, animated_coordinates
        )
        if triangulation is None:
            return []

        ax.tripcolor(
            triangulation,
            animated_values,
            shading="gouraud",
            cmap=cmap,
            vmin=value_min,
            vmax=value_max,
        )

        draw_contour_mesh_overlay(
            ax, elements, node_index, animated_coordinates, bc_nodes
        )
        ax.set_xlim(x_min, x_max)
        ax.set_ylim(y_min, y_max)
        ax.set_aspect("equal")
        ax.set_title(f"{title_prefix} {result_label} - faktor {amplitude:.3f}")
        ax.set_xlabel("X")
        ax.set_ylabel("Y")
        ax.grid(True, alpha=0.12)
        return []

    animation = FuncAnimation(
        fig,
        update,
        frames=max(int(frames), 2),
        interval=int(interval_ms),
        blit=False,
        repeat=True,
    )
    return save_animation_to_gif(fig, animation, interval_ms)


def plot_mesh(
    nodes,
    elements,
    node_index,
    coordinates,
    U,
    scale_factor,
    bc_nodes,
    critical_node_id=None,
    critical_element_id=None,
):
    fig, ax = plt.subplots(figsize=(9, 6))
    plot_coordinates = build_plot_coordinates(coordinates, U, scale_factor)

    for _, element in elements.iterrows():
        node_ids = [int(element[f"n{i}"]) for i in range(1, 5)]
        coords = np.array([coordinates[node_index[node_id]] for node_id in node_ids])
        coords_closed = np.vstack((coords, coords[0]))
        ax.plot(
            coords_closed[:, 0],
            coords_closed[:, 1],
            color="0.55",
            linestyle="--",
            linewidth=1.0,
        )

        coords_deformed = np.array(
            [plot_coordinates[node_index[node_id]] for node_id in node_ids]
        )
        coords_deformed_closed = np.vstack((coords_deformed, coords_deformed[0]))
        ax.plot(
            coords_deformed_closed[:, 0],
            coords_deformed_closed[:, 1],
            color="#005bbb",
            linewidth=1.8,
        )

    for _, node in nodes.iterrows():
        node_id = int(node["id"])
        idx = node_index[node_id]
        x = coordinates[idx, 0]
        y = coordinates[idx, 1]
        ax.text(x, y, str(node_id), fontsize=8, color="black")
        if node_id in bc_nodes:
            ax.scatter(x, y, s=30, c="#b22222", marker="s", zorder=3)

    if critical_node_id is not None and critical_node_id in node_index:
        critical_node_coordinates = plot_coordinates[node_index[critical_node_id]]
        ax.scatter(
            critical_node_coordinates[0],
            critical_node_coordinates[1],
            s=90,
            c="#d62828",
            marker="o",
            edgecolors="white",
            linewidths=1.1,
            zorder=5,
            label=f"Node kritis Umax: {critical_node_id}",
        )
        ax.annotate(
            f"Umax N{critical_node_id}",
            xy=(critical_node_coordinates[0], critical_node_coordinates[1]),
            xytext=(8, 8),
            textcoords="offset points",
            fontsize=8,
            color="#8b0000",
            bbox={"facecolor": "white", "alpha": 0.85, "edgecolor": "#d62828"},
        )

    if critical_element_id is not None:
        critical_element_rows = elements[elements["id"] == critical_element_id]
        if not critical_element_rows.empty:
            critical_element = critical_element_rows.iloc[0]
            critical_node_ids = [
                int(critical_element[f"n{i}"]) for i in range(1, 5)
            ]
            critical_centroid = np.array(
                [plot_coordinates[node_index[node_id]] for node_id in critical_node_ids]
            ).mean(axis=0)
            ax.scatter(
                critical_centroid[0],
                critical_centroid[1],
                s=170,
                c="#f77f00",
                marker="*",
                edgecolors="black",
                linewidths=0.8,
                zorder=5,
                label=f"Elemen kritis VM: {critical_element_id}",
            )
            ax.annotate(
                f"VMmax E{critical_element_id}",
                xy=(critical_centroid[0], critical_centroid[1]),
                xytext=(8, -16),
                textcoords="offset points",
                fontsize=8,
                color="#8c4f00",
                bbox={"facecolor": "white", "alpha": 0.85, "edgecolor": "#f77f00"},
            )

    ax.set_aspect("equal")
    ax.set_title("Mesh Asli (abu-abu) dan Deformasi (biru)")
    ax.set_xlabel("X")
    ax.set_ylabel("Y")
    ax.grid(True, alpha=0.2)
    if critical_node_id is not None or critical_element_id is not None:
        ax.legend(loc="upper right")
    fig.tight_layout()
    return fig


def plot_scalar_contour(
    elements,
    node_index,
    coordinates,
    U,
    scale_factor,
    bc_nodes,
    nodal_values,
    title_prefix,
    result_label,
    unit_label,
    use_deformed_shape,
    cmap,
    figure_size=(9, 6),
):
    plot_coordinates = build_plot_coordinates(
        coordinates,
        U if use_deformed_shape else None,
        scale_factor if use_deformed_shape else 0.0,
    )
    triangulation = build_quad_triangulation(elements, node_index, plot_coordinates)
    if triangulation is None:
        return None

    nodal_values = np.asarray(nodal_values, dtype=float)
    finite_values = nodal_values[np.isfinite(nodal_values)]
    if finite_values.size == 0:
        return None

    value_min = float(finite_values.min())
    value_max = float(finite_values.max())

    fig, ax = plt.subplots(figsize=figure_size)
    if np.isclose(value_min, value_max):
        delta = max(abs(value_min) * 1e-6, 1e-6)
        contour = ax.tripcolor(
            triangulation,
            nodal_values,
            shading="gouraud",
            cmap=cmap,
            vmin=value_min - delta,
            vmax=value_max + delta,
        )
    else:
        levels = np.linspace(value_min, value_max, 16)
        contour = ax.tricontourf(
            triangulation,
            nodal_values,
            levels=levels,
            cmap=cmap,
        )
        ax.tricontour(
            triangulation,
            nodal_values,
            levels=levels,
            colors="0.35",
            linewidths=0.35,
            alpha=0.25,
        )

    draw_contour_mesh_overlay(ax, elements, node_index, plot_coordinates, bc_nodes)
    geometry_label = "Bentuk terdeformasi" if use_deformed_shape else "Mesh asli"
    colorbar = fig.colorbar(contour, ax=ax, pad=0.02)
    colorbar.set_label(f"{result_label} ({unit_label})")
    ax.set_aspect("equal")
    ax.set_title(f"{title_prefix} {result_label} - {geometry_label}")
    ax.set_xlabel("X")
    ax.set_ylabel("Y")
    ax.grid(True, alpha=0.12)
    fig.tight_layout()
    return fig


def plot_stress_contour(
    elements,
    node_index,
    coordinates,
    U,
    scale_factor,
    bc_nodes,
    element_results,
    result_column,
    use_deformed_shape,
):
    if element_results.empty:
        return None

    nodal_values = average_element_result_to_nodes(
        elements, node_index, element_results, result_column
    )
    result_label = STRESS_CONTOUR_LABELS.get(result_column, result_column)
    return plot_scalar_contour(
        elements,
        node_index,
        coordinates,
        U,
        scale_factor,
        bc_nodes,
        nodal_values,
        "Kontur Tegangan",
        result_label,
        "MPa",
        use_deformed_shape,
        STRESS_CONTOUR_CMAP,
    )


def displacement_values_from_vector(U, component):
    displacement_matrix = np.asarray(U, dtype=float).reshape(-1, 2)
    component_index = 0 if component == "Ux" else 1
    return displacement_matrix[:, component_index]


def plot_displacement_contour(
    elements,
    node_index,
    coordinates,
    U,
    scale_factor,
    bc_nodes,
    component,
    use_deformed_shape,
):
    nodal_values = displacement_values_from_vector(U, component)
    return plot_scalar_contour(
        elements,
        node_index,
        coordinates,
        U,
        scale_factor,
        bc_nodes,
        nodal_values,
        "Kontur Deformasi",
        component,
        "satuan model",
        use_deformed_shape,
        DISPLACEMENT_CONTOUR_CMAP,
        figure_size=ANIMATION_FIGURE_SIZE,
    )


def stress_state_text(value):
    if value > 1e-12:
        return "tarik", "#2e7d32"
    if value < -1e-12:
        return "tekan", "#d62828"
    return "netral", "#444444"


def sketch_condition_text(s1, s2):
    if s1 >= 0.0 and s2 >= 0.0:
        return "KONDISI : TARIK DOMINAN", "(Principal Tension Zone)", "#7b2cbf"
    if s1 <= 0.0 and s2 <= 0.0:
        return "KONDISI : TEKAN DUA ARAH", "(Biaxial Compression)", "#1d4ed8"
    return "KONDISI : TARIK-TEKAN CAMPURAN", "(Mixed Principal State)", "#7b2cbf"


def crack_annotation_from_sigma1(s1, theta_crack_deg):
    if s1 > 1e-12:
        return (
            f"{float(theta_crack_deg):+.3f} deg",
            "Berpotensi terjadi retak tarik primer",
            "#d62828",
        )
    return ("-", "Tidak terjadi retak tarik primer", "#d62828")


def direction_vector_from_ccw_angle(angle_deg, length):
    theta = np.radians(angle_deg)
    return np.array([length * np.cos(theta), length * np.sin(theta)])


def ray_rectangle_intersection(angle_deg, half_width, half_height):
    direction = direction_vector_from_ccw_angle(angle_deg, 1.0)
    scales = []
    if abs(direction[0]) > 1e-12:
        scales.append(half_width / abs(direction[0]))
    if abs(direction[1]) > 1e-12:
        scales.append(half_height / abs(direction[1]))
    scale = min(scales) if scales else 0.0
    return direction * scale


def draw_angle_arc(ax, angle_ccw_deg, radius, color, center):
    visual_angle_deg = float(angle_ccw_deg)
    arc_angles = np.linspace(0.0, visual_angle_deg, 120)
    x_arc = center[0] + radius * np.cos(np.radians(arc_angles))
    y_arc = center[1] + radius * np.sin(np.radians(arc_angles))
    ax.plot(x_arc, y_arc, color=color, linewidth=2.0)

    if len(x_arc) >= 6:
        ax.annotate(
            "",
            xy=(x_arc[-1], y_arc[-1]),
            xytext=(x_arc[-6], y_arc[-6]),
            arrowprops={"arrowstyle": "->", "color": color, "linewidth": 2.0},
        )

    end_angle = visual_angle_deg
    text_radius = radius + 0.12
    text_x = center[0] + text_radius * np.cos(np.radians(end_angle))
    text_y = center[1] + text_radius * np.sin(np.radians(end_angle))
    text_on_right = text_x >= center[0]
    text_x += 0.03 if text_on_right else -0.03
    vertical_sign = 1.0 if angle_ccw_deg >= 0.0 else -1.0
    min_vertical_gap = 0.40 if abs(angle_ccw_deg) < 15.0 else 0.28
    text_y = center[1] + vertical_sign * max(
        abs(text_y - center[1]) + 0.08,
        min_vertical_gap,
    )
    x_min, x_max = ax.get_xlim()
    y_min, y_max = ax.get_ylim()
    x_padding = 0.16
    y_padding = 0.14
    if text_on_right:
        y_axis_clearance = center[0] + (1.18 if abs(angle_ccw_deg) < 15.0 else 1.05)
        text_x = max(text_x, y_axis_clearance)
        text_x = min(text_x, x_max - x_padding)
    else:
        text_x = max(text_x, x_min + x_padding)
    text_y = min(max(text_y, y_min + y_padding), y_max - y_padding)
    ax.text(
        text_x,
        text_y,
        f"{angle_ccw_deg:+.3f} deg",
        color=color,
        fontsize=11,
        fontweight="bold",
        ha="right" if text_on_right else "left",
        va="center",
    )


def wrap_sketch_text(text, width):
    return textwrap.fill(
        text,
        width=width,
        break_long_words=False,
        break_on_hyphens=False,
    )


def sketch_text_line_count(text):
    return text.count("\n") + 1


def make_text_area(text, color, fontsize, fontweight="bold"):
    return TextArea(
        text,
        textprops={
            "color": color,
            "fontsize": fontsize,
            "fontweight": fontweight,
            "family": "DejaVu Sans",
        },
    )


def add_centered_box(ax, child, anchor_x, anchor_y, loc):
    anchored_box = AnchoredOffsetbox(
        loc=loc,
        child=child,
        frameon=False,
        bbox_to_anchor=(anchor_x, anchor_y),
        bbox_transform=ax.transAxes,
        borderpad=0.0,
        pad=0.0,
    )
    ax.add_artist(anchored_box)


def _draw_principal_stress_sketch(ax, row, subplot_title):
    has_gauss_point = "gauss_point" in row
    element_id = int(row["element"]) if "element" in row else None
    s1 = float(row["s1"])
    s2 = float(row["s2"])
    theta1_display = principal_angle_display_deg(row["principal_angle_1_deg"])
    theta2_display = perpendicular_angle_from_principal_deg(theta1_display)
    theta1_color = "#2e7d32" if theta1_display >= 0.0 else "#d62828"

    s1_state, s1_color = stress_state_text(s1)
    s2_state, s2_color = stress_state_text(s2)
    condition_text, condition_subtext, condition_color = sketch_condition_text(s1, s2)
    theta_retak_text, crack_status_text, crack_status_color = crack_annotation_from_sigma1(
        s1, theta2_display
    )
    theta_retak_color = "#d62828"
    if theta_retak_text not in {"-", ""}:
        theta_retak_value = float(theta_retak_text.replace(" deg", ""))
        theta_retak_color = "#2e7d32" if theta_retak_value >= 0.0 else "#d62828"
    crack_status_text = wrap_sketch_text(crack_status_text, 23)
    sketch_y_offset = -0.08
    sketch_xlim = (-1.55, 1.80)
    sketch_ylim = (-2.80, 2.15)

    ax.set_xlim(*sketch_xlim)
    ax.set_ylim(*sketch_ylim)

    ax.add_patch(
        Rectangle(
            (0.01, 0.01),
            0.98,
            0.98,
            transform=ax.transAxes,
            fill=False,
            edgecolor="#5b8cc8",
            linewidth=1.8,
            zorder=20,
            clip_on=False,
        )
    )

    ax.add_patch(
        Rectangle(
            (-0.38, sketch_y_offset - 0.22),
            0.76,
            0.44,
            facecolor="#e8a9a2",
            edgecolor="0.2",
            linewidth=1.6,
            alpha=0.9,
        )
    )

    ax.plot(
        [-1.15, 1.15],
        [sketch_y_offset, sketch_y_offset],
        color="0.15",
        linewidth=1.4,
        dashes=(5, 5),
    )
    ax.plot(
        [0.0, 0.0],
        [sketch_y_offset - 1.05, sketch_y_offset + 1.05],
        color="0.15",
        linewidth=1.4,
        dashes=(5, 5),
    )
    ax.annotate(
        "",
        xy=(1.18, sketch_y_offset),
        xytext=(0.88, sketch_y_offset),
        arrowprops={"arrowstyle": "-|>", "color": "black", "linewidth": 1.5},
    )
    ax.annotate(
        "",
        xy=(0.0, sketch_y_offset + 1.08),
        xytext=(0.0, sketch_y_offset + 0.78),
        arrowprops={"arrowstyle": "-|>", "color": "black", "linewidth": 1.5},
    )
    ax.text(1.22, sketch_y_offset - 0.04, "x", fontsize=11)
    ax.text(-0.04, sketch_y_offset + 1.12, "y", fontsize=11)

    half_width = 0.38
    half_height = 0.22
    principal_axis_length = 1.20
    if s1 > 1e-12:
        crack_axis = direction_vector_from_ccw_angle(theta2_display, principal_axis_length)
        ax.plot(
            [-crack_axis[0], crack_axis[0]],
            [sketch_y_offset - crack_axis[1], sketch_y_offset + crack_axis[1]],
            color="#ff4d4d",
            linewidth=1.4,
            dashes=(5, 4),
            alpha=0.9,
        )

    sigma1_tail = direction_vector_from_ccw_angle(theta1_display + 180.0, 0.62)
    ax.plot(
        [0.0, sigma1_tail[0]],
        [sketch_y_offset, sketch_y_offset + sigma1_tail[1]],
        color="0.1",
        linewidth=1.4,
        dashes=(5, 5),
    )

    sigma1_start = ray_rectangle_intersection(theta1_display, half_width, half_height)
    sigma1_tip = sigma1_start + direction_vector_from_ccw_angle(theta1_display, 0.72)
    angle_arc_radius = max(0.68, np.linalg.norm(sigma1_tip) - 0.08)
    ax.annotate(
        "",
        xy=(sigma1_tip[0], sigma1_tip[1] + sketch_y_offset),
        xytext=(sigma1_start[0], sigma1_start[1] + sketch_y_offset),
        arrowprops={"arrowstyle": "->", "color": "#2e7d32", "linewidth": 2.2},
    )

    draw_angle_arc(
        ax,
        theta1_display,
        radius=angle_arc_radius,
        color=theta1_color,
        center=(0.0, sketch_y_offset),
    )

    if element_id is not None and not has_gauss_point:
        ax.text(
            0.045,
            0.945,
            f"{element_id}",
            transform=ax.transAxes,
            ha="center",
            va="center",
            fontsize=12,
            fontweight="bold",
            bbox={
                "facecolor": "white",
                "edgecolor": "black",
                "linewidth": 1.2,
                "boxstyle": "square,pad=0.42",
            },
        )

    header_line_1 = HPacker(
        children=[
            make_text_area(r"$\sigma_1$ = " + f"{s1:+.3f} MPa", "#111111", 9.5),
            make_text_area(f" ({s1_state})", s1_color, 9.5),
        ],
        align="center",
        pad=0,
        sep=2,
    )
    header_line_2 = HPacker(
        children=[
            make_text_area(r"$\sigma_2$ = " + f"{s2:+.3f} MPa", "#111111", 9.5),
            make_text_area(f" ({s2_state})", s2_color, 9.5),
        ],
        align="center",
        pad=0,
        sep=2,
    )
    header_box = VPacker(
        children=[header_line_1, header_line_2],
        align="center",
        pad=0,
        sep=2,
    )
    add_centered_box(ax, header_box, 0.54, 0.945, "upper center")

    theta_lines = VPacker(
        children=[
            make_text_area(
                r"$\theta\ (\sigma_1)$ = " + f"{theta1_display:+.3f}" + r"$^\circ$",
                theta1_color,
                11.5,
            ),
            make_text_area(
                r"$\theta_{retak}$ = "
                + (
                    theta_retak_text.replace(" deg", "") + r"$^\circ$"
                    if theta_retak_text != "-"
                    else "-"
                ),
                theta_retak_color,
                11.5,
            ),
        ],
        align="center",
        pad=0,
        sep=1,
    )
    condition_children = [
        theta_lines,
        make_text_area(condition_text, condition_color, 9.8),
        make_text_area(condition_subtext, condition_color, 9.5),
        make_text_area(crack_status_text, crack_status_color, 10.1),
    ]
    condition_box = VPacker(
        children=condition_children,
        align="center",
        pad=0,
        sep=4,
    )
    add_centered_box(ax, condition_box, 0.50, 0.075, "lower center")

    ax.set_xlim(*sketch_xlim)
    ax.set_ylim(*sketch_ylim)
    ax.set_aspect("equal", adjustable="box")
    ax.axis("off")
    if element_id is None or has_gauss_point:
        ax.set_title(subplot_title)


def plot_principal_stress_sketches(element_results):
    if element_results.empty:
        return None

    n_elements = len(element_results)
    ncols = min(3, max(1, n_elements))
    nrows = int(np.ceil(n_elements / ncols))
    fig, axes = plt.subplots(nrows, ncols, figsize=(7.0 * ncols, 5.4 * nrows))
    axes = np.atleast_1d(axes).flatten()

    for ax, (_, row) in zip(axes, element_results.iterrows()):
        _draw_principal_stress_sketch(ax, row, f"Elemen {int(row['element'])}")

    for ax in axes[n_elements:]:
        ax.axis("off")

    fig.suptitle("Sketsa Tegangan Utama per Elemen", fontsize=14)
    fig.tight_layout(rect=[0.0, 0.0, 1.0, 0.97])
    return fig


def plot_principal_stress_sketches_for_element_gauss(gauss_results, element_id):
    element_rows = gauss_results[gauss_results["element"] == element_id].sort_values(
        "gauss_point"
    )
    if element_rows.empty:
        return None

    n_points = len(element_rows)
    ncols = min(2, max(1, n_points))
    nrows = int(np.ceil(n_points / ncols))
    fig, axes = plt.subplots(nrows, ncols, figsize=(7.0 * ncols, 5.4 * nrows))
    axes = np.atleast_1d(axes).flatten()

    for ax, (_, row) in zip(axes, element_rows.iterrows()):
        _draw_principal_stress_sketch(
            ax,
            row,
            (
                f"Elemen {element_id} - GP {int(row['gauss_point'])}\n"
                f"xi={float(row['xi']):.4f}, eta={float(row['eta']):.4f}"
            ),
        )

    for ax in axes[n_points:]:
        ax.axis("off")

    fig.suptitle(
        f"Sketsa Tegangan Utama Titik Gauss Elemen {element_id}", fontsize=14
    )
    fig.tight_layout(rect=[0.0, 0.0, 1.0, 0.97])
    return fig


def _draw_mohr_circle(ax, row, subplot_title):
    sx = float(row["sx"])
    sy = float(row["sy"])
    txy = float(row["txy"])
    s1 = float(row["s1"])
    s2 = float(row["s2"])
    center = 0.5 * (sx + sy)
    radius = np.sqrt(((sx - sy) / 2.0) ** 2 + txy**2)

    theta = np.linspace(0.0, 2.0 * np.pi, 400)
    sigma_circle = center + radius * np.cos(theta)
    tau_circle = radius * np.sin(theta)

    ax.plot(sigma_circle, tau_circle, color="#005bbb", linewidth=2.0)
    ax.axhline(0.0, color="0.25", linewidth=0.8)
    ax.axvline(0.0, color="0.25", linewidth=0.8)

    ax.scatter([sx, sy], [txy, -txy], color="#b22222", s=35, zorder=3)
    ax.scatter([s1, s2, center], [0.0, 0.0, 0.0], color="#2e8b57", s=30, zorder=3)

    ax.annotate("A", (sx, txy), textcoords="offset points", xytext=(6, 6), fontsize=9)
    ax.annotate("B", (sy, -txy), textcoords="offset points", xytext=(6, -12), fontsize=9)
    ax.annotate("sigma1", (s1, 0.0), textcoords="offset points", xytext=(6, 6), fontsize=9)
    ax.annotate("sigma2", (s2, 0.0), textcoords="offset points", xytext=(6, 6), fontsize=9)
    ax.annotate("C", (center, 0.0), textcoords="offset points", xytext=(6, 6), fontsize=9)

    margin = max(radius * 0.2, 1e-9)
    x_min = min(sx, sy, s1, s2, center) - margin
    x_max = max(sx, sy, s1, s2, center) + margin
    y_lim = radius + margin
    if y_lim <= 0.0:
        y_lim = 1.0

    ax.set_xlim(x_min, x_max)
    ax.set_ylim(-y_lim, y_lim)
    ax.set_aspect("equal", adjustable="box")
    ax.grid(True, alpha=0.25)
    ax.set_title(subplot_title)
    ax.set_xlabel("Tegangan Normal, sigma (MPa)")
    ax.set_ylabel("Tegangan Geser, tau (MPa)")

    summary = (
        f"C={center:.3g}\n"
        f"R={radius:.3g}\n"
        f"sigma1={s1:.3g}\n"
        f"sigma2={s2:.3g}"
    )
    ax.text(
        0.02,
        0.98,
        summary,
        transform=ax.transAxes,
        va="top",
        ha="left",
        fontsize=8,
        bbox={"facecolor": "white", "alpha": 0.85, "edgecolor": "0.7"},
    )


def plot_mohr_circles(element_results):
    if element_results.empty:
        return None

    n_elements = len(element_results)
    ncols = min(3, max(1, n_elements))
    nrows = int(np.ceil(n_elements / ncols))
    fig, axes = plt.subplots(nrows, ncols, figsize=(5.5 * ncols, 4.5 * nrows))
    axes = np.atleast_1d(axes).flatten()

    for ax, (_, row) in zip(axes, element_results.iterrows()):
        _draw_mohr_circle(ax, row, f"Elemen {int(row['element'])}")

    for ax in axes[n_elements:]:
        ax.axis("off")

    fig.suptitle("Lingkaran Mohr per Elemen", fontsize=14)
    fig.tight_layout()
    return fig


def plot_mohr_circles_for_element_gauss(gauss_results, element_id):
    element_rows = gauss_results[gauss_results["element"] == element_id].sort_values(
        "gauss_point"
    )
    if element_rows.empty:
        return None

    n_points = len(element_rows)
    ncols = min(2, max(1, n_points))
    nrows = int(np.ceil(n_points / ncols))
    fig, axes = plt.subplots(nrows, ncols, figsize=(5.5 * ncols, 4.5 * nrows))
    axes = np.atleast_1d(axes).flatten()

    for ax, (_, row) in zip(axes, element_rows.iterrows()):
        _draw_mohr_circle(
            ax,
            row,
            (
                f"Elemen {element_id} - GP {int(row['gauss_point'])}\n"
                f"xi={float(row['xi']):.4f}, eta={float(row['eta']):.4f}"
            ),
        )

    for ax in axes[n_points:]:
        ax.axis("off")

    fig.suptitle(f"Lingkaran Mohr Titik Gauss Elemen {element_id}", fontsize=14)
    fig.tight_layout()
    return fig


def build_template_workbook():
    nodes = pd.DataFrame(
        [
            [1, 0.0, 0.0],
            [2, 100.0, 0.0],
            [3, 200.0, 0.0],
            [4, 300.0, 0.0],
            [5, 300.0, 50.0],
            [6, 200.0, 50.0],
            [7, 100.0, 50.0],
            [8, 0.0, 50.0],
            [9, 300.0, 100.0],
            [10, 200.0, 100.0],
            [11, 100.0, 100.0],
            [12, 0.0, 100.0],
            [13, 300.0, 150.0],
            [14, 200.0, 150.0],
            [15, 100.0, 150.0],
            [16, 0.0, 150.0],
        ],
        columns=["id", "x", "y"],
    )
    elements = pd.DataFrame(
        [
            [1, 1, 2, 7, 8, 2.10e5, 0.31, 10.0],
            [2, 2, 3, 6, 7, 2.10e5, 0.31, 10.0],
            [3, 3, 4, 5, 6, 2.10e5, 0.31, 10.0],
            [4, 8, 7, 11, 12, 2.10e5, 0.31, 10.0],
            [5, 6, 5, 9, 10, 2.10e5, 0.31, 10.0],
            [6, 12, 11, 15, 16, 2.10e5, 0.31, 10.0],
            [7, 11, 10, 14, 15, 2.10e5, 0.31, 10.0],
            [8, 10, 9, 13, 14, 2.10e5, 0.31, 10.0],
        ],
        columns=["id", "n1", "n2", "n3", "n4", "E", "nu", "t"],
    )
    loads = pd.DataFrame([[13, 0.0, -4000.0]], columns=["node", "Fx", "Fy"])
    bc = pd.DataFrame(
        [
            [1, 1, 1, 0.0, 0.0],
            [8, 1, 1, 0.0, 0.0],
            [12, 1, 1, 0.0, 0.0],
            [16, 1, 1, 0.0, 0.0],
        ],
        columns=["node", "ux", "uy", "ux_val", "uy_val"],
    )
    edge_loads = pd.DataFrame(
        [
            [6, 3, 0.0, -5.0],
            [7, 3, 0.0, -5.0],
            [8, 3, 0.0, -5.0],
        ],
        columns=["element", "edge", "qx", "qy"],
    )
    info = pd.DataFrame(
        [
            ["Nodes", "id, x, y", "Koordinat node, gunakan satuan yang konsisten."],
            [
                "Elements",
                "id, n1, n2, n3, n4, E, nu, t",
                "Urutan node harus berlawanan arah jarum jam.",
            ],
            ["Loads", "node, Fx, Fy", "Beban terpusat nodal."],
            [
                "BC",
                "node, ux, uy, ux_val, uy_val",
                "1 = dikunci, 0 = bebas. Nilai default displacement = 0.",
            ],
            [
                "EdgeLoads",
                "element, edge, qx, qy",
                "Beban garis seragam per panjang pada edge 1-4 dalam koordinat global.",
            ],
        ],
        columns=["Sheet", "Kolom", "Keterangan"],
    )

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        info.to_excel(writer, sheet_name="Info", index=False)
        nodes.to_excel(writer, sheet_name="Nodes", index=False)
        elements.to_excel(writer, sheet_name="Elements", index=False)
        loads.to_excel(writer, sheet_name="Loads", index=False)
        bc.to_excel(writer, sheet_name="BC", index=False)
        edge_loads.to_excel(writer, sheet_name="EdgeLoads", index=False)
    output.seek(0)
    return output.getvalue()


def build_fallback_ulm_logo_svg():
    return """
    <svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 240 240">
      <defs>
        <linearGradient id="goldGlow" x1="0" y1="0" x2="0" y2="1">
          <stop offset="0%" stop-color="#fff27a"/>
          <stop offset="100%" stop-color="#ffd100"/>
        </linearGradient>
      </defs>
      <polygon
        points="120,12 192,48 205,126 165,220 75,220 35,126 48,48"
        fill="url(#goldGlow)"
        stroke="#d62828"
        stroke-width="10"
        stroke-linejoin="round"
      />
      <circle cx="120" cy="115" r="63" fill="#ffffff" stroke="#d62828" stroke-width="12"/>
      <circle cx="120" cy="115" r="47" fill="#fff6cf" stroke="#d62828" stroke-width="4"/>
      <path d="M72 112 C88 82, 104 74, 120 86 C104 92, 94 102, 82 120 Z" fill="#20242c"/>
      <path d="M168 112 C152 82, 136 74, 120 86 C136 92, 146 102, 158 120 Z" fill="#20242c"/>
      <path d="M108 95 L120 138 L132 95 L120 104 Z" fill="#20242c"/>
      <circle cx="120" cy="120" r="9" fill="#d62828"/>
      <text x="120" y="58" text-anchor="middle" font-size="18" font-family="Arial, sans-serif" font-weight="700" fill="#ffffff">UNIVERSITAS</text>
      <text x="120" y="188" text-anchor="middle" font-size="16" font-family="Arial, sans-serif" font-weight="700" fill="#ffffff">LAMBUNG MANGKURAT</text>
      <text x="120" y="162" text-anchor="middle" font-size="18" font-family="Arial, sans-serif" font-weight="800" fill="#d62828">ULM</text>
    </svg>
    """


def image_path_to_data_uri(image_path):
    suffix = image_path.suffix.lower()
    mime_type_map = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".svg": "image/svg+xml",
        ".webp": "image/webp",
        ".gif": "image/gif",
    }
    mime_type = mime_type_map.get(suffix, "application/octet-stream")
    encoded_bytes = base64.b64encode(image_path.read_bytes()).decode("ascii")
    return f"data:{mime_type};base64,{encoded_bytes}"


def get_header_logo_data_uri():
    candidate_names = [
        "Logo_ULM.png",
        "Logo_ULM.svg",
        "logo_ulm.png",
        "logo_ulm.svg",
        "ulm_logo.png",
        "ulm_logo.svg",
        "logo_ftulm.png",
        "logo_ftulm.svg",
        "logo.png",
        "logo.svg",
    ]
    search_roots = [Path(__file__).resolve().parent, Path.cwd()]
    for root_path in search_roots:
        available_files = {
            path.name.lower(): path for path in root_path.iterdir() if path.is_file()
        }
        for candidate_name in candidate_names:
            candidate_path = root_path / candidate_name
            if candidate_path.exists():
                return image_path_to_data_uri(candidate_path)

            # Streamlit Cloud runs on Linux, so "Logo_ULM.png" and "logo_ulm.png"
            # are different files. Fall back to a case-insensitive directory lookup.
            matched_path = available_files.get(candidate_name.lower())
            if matched_path is not None:
                return image_path_to_data_uri(matched_path)

    fallback_svg = build_fallback_ulm_logo_svg().strip().encode("utf-8")
    return f"data:image/svg+xml;base64,{base64.b64encode(fallback_svg).decode('ascii')}"


def render_application_header():
    logo_data_uri = get_header_logo_data_uri()
    st.markdown(
        f"""
        <style>
        .plane2d-hero {{
            background:
                radial-gradient(circle at top left, rgba(214, 40, 40, 0.10), transparent 28%),
                linear-gradient(145deg, #0c111b 0%, #111827 100%);
            border: 1px solid rgba(255, 255, 255, 0.08);
            border-radius: 24px;
            padding: 1.8rem 2rem;
            margin: 0.2rem 0 1.2rem 0;
            box-shadow: 0 22px 48px rgba(0, 0, 0, 0.22);
        }}
        .plane2d-hero-row {{
            display: flex;
            align-items: center;
            gap: 1.4rem;
        }}
        .plane2d-logo-shell {{
            flex: 0 0 168px;
            display: flex;
            justify-content: center;
            align-items: center;
        }}
        .plane2d-logo-shell img {{
            width: 150px;
            max-width: 100%;
            display: block;
            filter: drop-shadow(0 10px 18px rgba(0, 0, 0, 0.28));
        }}
        .plane2d-copy {{
            min-width: 0;
        }}
        .plane2d-kicker {{
            color: #f3c969;
            letter-spacing: 0.12em;
            text-transform: uppercase;
            font-size: 0.82rem;
            font-weight: 700;
            margin-bottom: 0.25rem;
        }}
        .plane2d-title {{
            color: #f8fafc;
            font-size: clamp(2.0rem, 3.6vw, 3.1rem);
            line-height: 1.08;
            font-weight: 800;
            margin: 0;
        }}
        .plane2d-subtitle {{
            color: #f8fafc;
            font-size: clamp(1.15rem, 2.1vw, 1.55rem);
            line-height: 1.25;
            font-weight: 700;
            margin-top: 0.2rem;
        }}
        .plane2d-meta {{
            color: #e5e7eb;
            font-size: 1.02rem;
            line-height: 1.45;
            font-weight: 600;
            margin-top: 0.9rem;
        }}
        .plane2d-divider {{
            height: 1px;
            margin: 1.35rem 0 1.2rem 0;
            background: linear-gradient(90deg, transparent, rgba(244, 244, 245, 0.85), transparent);
        }}
        .plane2d-disclaimer-title {{
            color: #f8fafc;
            font-size: 1.2rem;
            font-weight: 800;
            margin-bottom: 0.75rem;
        }}
        .plane2d-disclaimer-box {{
            background: linear-gradient(145deg, rgba(87, 91, 19, 0.95), rgba(75, 78, 14, 0.98));
            color: #f7f5cf;
            border: 1px solid rgba(248, 250, 252, 0.08);
            border-radius: 18px;
            padding: 1.15rem 1.25rem 1rem 1.4rem;
        }}
        .plane2d-disclaimer-box ul {{
            margin: 0;
            padding-left: 1.2rem;
        }}
        .plane2d-disclaimer-box li {{
            margin: 0 0 0.95rem 0;
            line-height: 1.6;
            font-size: 1rem;
        }}
        .plane2d-disclaimer-box li:last-child {{
            margin-bottom: 0;
        }}
        @media (max-width: 900px) {{
            .plane2d-hero {{
                padding: 1.25rem 1rem;
            }}
            .plane2d-hero-row {{
                flex-direction: column;
                align-items: flex-start;
            }}
            .plane2d-logo-shell {{
                flex-basis: auto;
            }}
            .plane2d-logo-shell img {{
                width: 124px;
            }}
        }}
        </style>
        <section class="plane2d-hero">
            <div class="plane2d-hero-row">
                <div class="plane2d-logo-shell">
                    <img src="{logo_data_uri}" alt="Logo Universitas Lambung Mangkurat" />
                </div>
                <div class="plane2d-copy">
                    <div class="plane2d-kicker">Finite Element Method Application</div>
                    <div class="plane2d-title">Analisis FEM Plane Stress & Strain 2D</div>
                    <div class="plane2d-subtitle">(Elemen Segiempat 4 Node)</div>
                    <div class="plane2d-meta">Pengembang: Ir. Darmansyah Tjitradi, MT., IPU</div>
                    <div class="plane2d-meta">Fakultas Teknik Universitas Lambung Mangkurat</div>
                </div>
            </div>
            <div class="plane2d-divider"></div>
            <div class="plane2d-disclaimer-title">Disclaimer:</div>
            <div class="plane2d-disclaimer-box">
                <ul>
                    <li>Aplikasi ini dikembangkan sebagai alat bantu pembelajaran dan penelitian dalam analisis struktur menggunakan Metode Elemen Hingga (Finite Element Method).</li>
                    <li>Hasil analisis harus diverifikasi lebih lanjut oleh insinyur profesional yang kompeten sebelum digunakan dalam perancangan struktur.</li>
                    <li>Pengguna bertanggung jawab sepenuhnya atas interpretasi dan penggunaan hasil analisis yang diperoleh dari aplikasi ini.</li>
                </ul>
            </div>
        </section>
        """,
        unsafe_allow_html=True,
    )


def write_local_template(template_bytes):
    target = Path(__file__).resolve().with_name(TEMPLATE_FILENAME)
    if target.exists():
        return target
    try:
        target.write_bytes(template_bytes)
        return target
    except PermissionError:
        fallback = Path.cwd() / TEMPLATE_FILENAME
        if fallback.exists():
            return fallback
        fallback.write_bytes(template_bytes)
        return fallback


render_application_header()
st.caption(
    "Formulasi memakai elemen isoparametrik bilinear dengan integrasi Gauss 2x2. "
    "Input dibaca dari Excel dan urutan node elemen harus berlawanan arah jarum jam."
)
st.caption(
    "Semua output tegangan ditampilkan dalam MPa, dengan asumsi input material "
    "tegangan memakai satuan kg/cm2 dan konversi 1 kg/cm2 = 0.1 MPa."
)

analysis_mode = st.selectbox("Tipe analisis", ["Plane Stress", "Plane Strain"])
default_E = st.number_input(
    "Modulus elastisitas default E (kg/cm2)", value=2.10e5, format="%.6g"
)
default_nu = st.number_input(
    "Poisson ratio default", min_value=0.0, max_value=0.4999, value=0.31, format="%.4f"
)
default_t = st.number_input("Tebal default elemen (cm)", min_value=1e-9, value=10.0)
concrete_fc = st.number_input(
    "Kuat tekan beton fc' untuk Drucker-Prager (kg/cm2)",
    min_value=1e-9,
    value=25.0,
    format="%.6g",
)
concrete_ft = st.number_input(
    "Kuat tarik beton ft untuk Drucker-Prager (kg/cm2)",
    min_value=1e-9,
    value=2.5,
    format="%.6g",
)
st.caption(
    "Evaluasi Drucker-Prager memakai asumsi post-processing plane stress dengan "
    "kalibrasi dari `fc'` dan `ft`, serta konvensi tegangan tarik bernilai positif."
)

template_bytes = build_template_workbook()
local_template_path = write_local_template(template_bytes)
local_example_path = Path(__file__).resolve().with_name(EXAMPLE_FILENAME)
st.download_button(
    "Unduh template Excel contoh",
    data=template_bytes,
    file_name="template_plane_stress_quad4.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
st.info(
    f"Jika tombol download tidak berjalan, template lokal otomatis disimpan di: `{local_template_path}`"
)
if local_example_path.exists():
    st.info(
        f"Contoh file Excel siap pakai juga tersedia di folder project: `{local_example_path}`"
    )

with st.expander("Format input Excel", expanded=False):
    st.write("Sheet wajib: `Nodes`, `Elements`, `BC`.")
    st.write("Sheet opsional: `Loads`, `EdgeLoads`.")
    st.write("Kolom minimum:")
    st.write("- `Nodes`: `id`, `x`, `y`")
    st.write("- `Elements`: `id`, `n1`, `n2`, `n3`, `n4`")
    st.write("- `BC`: `node`, `ux`, `uy`")
    st.write("- `Loads`: `node`, `Fx`, `Fy`")
    st.write("- `EdgeLoads`: `element`, `edge`, `qx`, `qy`")
    st.write(
        "Jika kolom `E`, `nu`, atau `t` tidak ada pada `Elements`, program memakai nilai default dari panel input."
    )
    st.write(
        "Asumsi satuan saat ini: `E`, `fc'`, dan `ft` diinput dalam `kg/cm2`, sedangkan semua output tegangan ditampilkan dalam `MPa`."
    )

uploaded_file = st.file_uploader("Upload file Excel", type=["xlsx"])

if uploaded_file:
    try:
        xls = pd.ExcelFile(uploaded_file)
        nodes = read_required_sheet(xls, "Nodes", ["id", "x", "y"]).copy()
        elements = read_required_sheet(
            xls, "Elements", ["id", "n1", "n2", "n3", "n4"]
        ).copy()
        bc = read_required_sheet(xls, "BC", ["node", "ux", "uy"]).copy()
        loads = read_optional_sheet(xls, "Loads", ["node", "Fx", "Fy"]).copy()
        edge_loads = read_optional_sheet(
            xls, "EdgeLoads", ["element", "edge", "qx", "qy"]
        ).copy()

        nodes["id"] = nodes["id"].astype(int)
        elements["id"] = elements["id"].astype(int)
        for column in ["n1", "n2", "n3", "n4"]:
            elements[column] = elements[column].astype(int)
        bc["node"] = bc["node"].astype(int)
        if not loads.empty:
            loads["node"] = loads["node"].astype(int)
        if not edge_loads.empty:
            edge_loads["element"] = edge_loads["element"].astype(int)
            edge_loads["edge"] = edge_loads["edge"].astype(int)

        if nodes["id"].duplicated().any():
            raise ValueError("ID node harus unik.")
        if elements["id"].duplicated().any():
            raise ValueError("ID elemen harus unik.")

        node_ids = nodes["id"].tolist()
        node_index = {node_id: idx for idx, node_id in enumerate(node_ids)}
        coordinates = nodes[["x", "y"]].to_numpy(dtype=float)

        for _, element in elements.iterrows():
            element_id = int(element["id"])
            element_nodes = [int(element[f"n{i}"]) for i in range(1, 5)]
            missing_nodes = [node_id for node_id in element_nodes if node_id not in node_index]
            if missing_nodes:
                raise ValueError(
                    f"Elemen {element_id} memakai node yang tidak ada: {missing_nodes}"
                )
            coords = np.array([coordinates[node_index[node_id]] for node_id in element_nodes])
            if signed_polygon_area(coords) <= 0.0:
                raise ValueError(
                    f"Elemen {element_id} memiliki urutan node tidak valid. Gunakan urutan berlawanan arah jarum jam."
                )

        if not loads.empty:
            missing_load_nodes = [
                int(node_id) for node_id in loads["node"] if int(node_id) not in node_index
            ]
            if missing_load_nodes:
                raise ValueError(
                    f"Ada node pada sheet Loads yang tidak ditemukan: {sorted(set(missing_load_nodes))}"
                )

        dof = 2 * len(nodes)
        K = np.zeros((dof, dof))
        F = np.zeros(dof)
        element_lookup = {}
        element_debug_data = []
        global_labels = global_dof_labels(node_ids)

        for _, element in elements.iterrows():
            element_id = int(element["id"])
            element_nodes = [int(element[f"n{i}"]) for i in range(1, 5)]
            coords = np.array([coordinates[node_index[node_id]] for node_id in element_nodes])
            E = row_value(element, "E", default_E)
            nu = row_value(element, "nu", default_nu)
            thickness = row_value(element, "t", default_t)

            if E <= 0.0:
                raise ValueError(f"Nilai E elemen {element_id} harus lebih besar dari nol.")
            if thickness <= 0.0:
                raise ValueError(
                    f"Nilai tebal elemen {element_id} harus lebih besar dari nol."
                )
            if not (0.0 <= nu < 0.5):
                raise ValueError(
                    f"Poisson ratio elemen {element_id} harus berada pada rentang 0 sampai < 0.5."
                )

            D = constitutive_matrix(E, nu, analysis_mode)
            element_matrices = quad4_element_matrices(coords, D, thickness)
            Ke = element_matrices["ke"]
            dof_map = element_dof_map(element_nodes, node_index)
            Kg_element = assemble_element_global_matrix(Ke, dof_map, dof)

            K[np.ix_(dof_map, dof_map)] += Ke

            element_lookup[element_id] = {
                "coords": coords,
                "dof_map": dof_map,
                "node_ids": element_nodes,
            }
            element_debug_data.append(
                {
                    "element_id": element_id,
                    "node_ids": element_nodes,
                    "coords": coords,
                    "D": D,
                    "dof_map": dof_map,
                    "local_labels": element_dof_labels(element_nodes),
                    "gauss_data": element_matrices["gauss_data"],
                    "B_total": element_matrices["B_total"],
                    "ke": Ke,
                    "KG_element": Kg_element,
                }
            )

        for _, load in loads.iterrows():
            idx = node_index[int(load["node"])]
            F[2 * idx] += row_value(load, "Fx", 0.0)
            F[2 * idx + 1] += row_value(load, "Fy", 0.0)

        for _, edge_load in edge_loads.iterrows():
            element_id = int(edge_load["element"])
            if element_id not in element_lookup:
                raise ValueError(
                    f"Edge load mengacu ke elemen {element_id}, tetapi elemen itu tidak ada."
                )
            edge = int(edge_load["edge"])
            fe = quad4_edge_load(
                element_lookup[element_id]["coords"],
                edge,
                row_value(edge_load, "qx", 0.0),
                row_value(edge_load, "qy", 0.0),
            )
            dof_map = element_lookup[element_id]["dof_map"]
            for i in range(8):
                F[dof_map[i]] += fe[i]

        prescribed_dofs = {}
        constrained_nodes = set()
        for _, constraint in bc.iterrows():
            node_id = int(constraint["node"])
            if node_id not in node_index:
                raise ValueError(
                    f"Boundary condition mengacu ke node {node_id}, tetapi node itu tidak ada."
                )
            constrained_nodes.add(node_id)
            idx = node_index[node_id]

            if int(constraint["ux"]) == 1:
                dof_id = 2 * idx
                ux_value = row_value(constraint, "ux_val", 0.0)
                if dof_id in prescribed_dofs and not np.isclose(prescribed_dofs[dof_id], ux_value):
                    raise ValueError(f"DOF ux pada node {node_id} memiliki nilai constraint ganda.")
                prescribed_dofs[dof_id] = ux_value

            if int(constraint["uy"]) == 1:
                dof_id = 2 * idx + 1
                uy_value = row_value(constraint, "uy_val", 0.0)
                if dof_id in prescribed_dofs and not np.isclose(prescribed_dofs[dof_id], uy_value):
                    raise ValueError(f"DOF uy pada node {node_id} memiliki nilai constraint ganda.")
                prescribed_dofs[dof_id] = uy_value

        U, reactions = solve_system(K, F, prescribed_dofs)

        displacement_df, gauss_results_df, element_results_df = create_results(
            nodes,
            elements,
            node_index,
            coordinates,
            analysis_mode,
            default_E,
            default_nu,
            default_t,
            concrete_fc,
            concrete_ft,
            U,
        )
        gauss_results_mpa_df = convert_stress_dataframe_to_mpa(gauss_results_df)
        element_results_mpa_df = convert_stress_dataframe_to_mpa(element_results_df)
        gauss_results_display_df = add_counterclockwise_principal_angle_columns(
            gauss_results_mpa_df
        )
        element_results_display_df = add_counterclockwise_principal_angle_columns(
            element_results_mpa_df
        )

        reaction_rows = []
        for _, node in nodes.iterrows():
            node_id = int(node["id"])
            idx = node_index[node_id]
            reaction_rows.append(
                {
                    "node": node_id,
                    "Rx": reactions[2 * idx],
                    "Ry": reactions[2 * idx + 1],
                }
            )
        reaction_df = pd.DataFrame(reaction_rows)
        constrained_reaction_df = reaction_df[
            reaction_df["node"].isin(sorted(constrained_nodes))
        ].copy()
        constrained_reaction_display_df = constrained_reaction_df.merge(
            displacement_df[["node", "x", "y"]],
            on="node",
            how="left",
        )
        constrained_reaction_display_df["R_mag"] = np.hypot(
            constrained_reaction_display_df["Rx"],
            constrained_reaction_display_df["Ry"],
        )
        constrained_reaction_display_df = constrained_reaction_display_df[
            ["node", "x", "y", "Rx", "Ry", "R_mag"]
        ]

        support_reaction_detail_df = build_support_reaction_detail_dataframe(
            reactions, prescribed_dofs, node_ids
        )
        equilibrium_df, equilibrium_ok = build_equilibrium_dataframe(
            F, reactions, prescribed_dofs
        )
        model_summary_df = build_model_summary_dataframe(
            analysis_mode,
            nodes,
            elements,
            loads,
            bc,
            edge_loads,
            default_E,
            default_nu,
            default_t,
            concrete_fc,
            concrete_ft,
            dof,
            prescribed_dofs,
        )
        executive_summary_df = build_executive_summary_dataframe(
            displacement_df,
            element_results_display_df,
            gauss_results_display_df,
            support_reaction_detail_df,
        )
        nodal_hotspots_df = build_nodal_hotspots_dataframe(displacement_df)
        reaction_hotspots_df = build_support_hotspots_dataframe(
            constrained_reaction_df,
            displacement_df,
        )
        element_vm_hotspots_df = build_element_hotspot_dataframe(
            element_results_display_df,
            "von_mises",
        )
        element_yield_hotspots_df = build_element_hotspot_dataframe(
            element_results_display_df,
            "yield_function",
        )
        gauss_vm_hotspots_df = build_gauss_hotspot_dataframe(
            gauss_results_display_df,
            "von_mises",
        )
        gauss_yield_hotspots_df = build_gauss_hotspot_dataframe(
            gauss_results_display_df,
            "yield_function",
        )

        critical_displacement_row = select_extreme_row(displacement_df, "U_mag")
        critical_element_row = select_extreme_row(
            element_results_display_df, "von_mises"
        )
        critical_gauss_row = select_extreme_row(gauss_results_display_df, "von_mises")
        critical_node_id = (
            int(critical_displacement_row["node"])
            if critical_displacement_row is not None
            else None
        )
        critical_element_id = (
            int(critical_element_row["element"])
            if critical_element_row is not None
            else None
        )

        max_disp = displacement_df["U_mag"].max() if not displacement_df.empty else 0.0
        max_vm_gauss = (
            gauss_results_mpa_df["von_mises"].max()
            if not gauss_results_mpa_df.empty
            else 0.0
        )
        yield_point_count = (
            int((gauss_results_display_df["yield_state"] == "Yield").sum())
            if "yield_state" in gauss_results_display_df.columns
            else 0
        )

        export_notes_rows = [
            {
                "Catatan": (
                    "Semua kolom tegangan pada workbook hasil diekspor dalam MPa. "
                    "Displacement, beban, reaksi, dan matriks kekakuan tetap mengikuti satuan model input."
                )
            },
            {
                "Catatan": (
                    "Hotspot elemen memakai hasil berbasis `B_total`, sedangkan hotspot titik Gauss "
                    "lebih konservatif karena dievaluasi per titik integrasi."
                )
            },
        ]
        result_sheet_specs = [
            {"sheet_name": "ExecutiveSummary", "dataframe": executive_summary_df},
            {"sheet_name": "ModelInfo", "dataframe": model_summary_df},
            {"sheet_name": "Equilibrium", "dataframe": equilibrium_df},
            {"sheet_name": "Hotspot_Nodes", "dataframe": nodal_hotspots_df},
            {"sheet_name": "Hotspot_Reactions", "dataframe": reaction_hotspots_df},
            {"sheet_name": "Hotspot_Elem_VM", "dataframe": element_vm_hotspots_df},
            {"sheet_name": "Hotspot_Elem_Yield", "dataframe": element_yield_hotspots_df},
            {"sheet_name": "Hotspot_GP_VM", "dataframe": gauss_vm_hotspots_df},
            {"sheet_name": "Hotspot_GP_Yield", "dataframe": gauss_yield_hotspots_df},
            {"sheet_name": "Displacement_Nodes", "dataframe": displacement_df},
            {
                "sheet_name": "Support_Reactions",
                "dataframe": constrained_reaction_display_df,
            },
            {
                "sheet_name": "Support_Reaction_DOF",
                "dataframe": support_reaction_detail_df,
            },
            {"sheet_name": "Element_Results", "dataframe": element_results_display_df},
            {"sheet_name": "Gauss_Results", "dataframe": gauss_results_display_df},
            {
                "sheet_name": "Vector_P",
                "dataframe": vector_to_dataframe(F, global_labels, "P"),
            },
            {
                "sheet_name": "Vector_d",
                "dataframe": vector_to_dataframe(U, global_labels, "d"),
            },
            {
                "sheet_name": "Vector_R",
                "dataframe": vector_to_dataframe(reactions, global_labels, "R"),
            },
            {"sheet_name": "Input_Nodes", "dataframe": nodes},
            {"sheet_name": "Input_Elements", "dataframe": elements},
            {"sheet_name": "Input_Loads", "dataframe": loads},
            {"sheet_name": "Input_BC", "dataframe": bc},
            {"sheet_name": "Input_EdgeLoads", "dataframe": edge_loads},
        ]
        if K.size <= MAX_EXPORT_MATRIX_CELLS:
            result_sheet_specs.append(
                {
                    "sheet_name": "Global_KG",
                    "dataframe": matrix_to_dataframe(
                        K,
                        global_labels,
                        global_labels,
                    ),
                    "include_index": True,
                }
            )
        else:
            export_notes_rows.append(
                {
                    "Catatan": (
                        f"Sheet `Global_KG` tidak diekspor karena ukurannya {K.shape[0]}x{K.shape[1]} "
                        f"({K.size} sel), melebihi batas otomatis {MAX_EXPORT_MATRIX_CELLS} sel."
                    )
                }
            )
        if critical_gauss_row is not None:
            export_notes_rows.append(
                {
                    "Catatan": (
                        "Von Mises maksimum titik Gauss berada pada "
                        f"elemen {int(critical_gauss_row['element'])}, GP {int(critical_gauss_row['gauss_point'])}."
                    )
                }
            )
        result_sheet_specs.append(
            {"sheet_name": "ExportNotes", "dataframe": pd.DataFrame(export_notes_rows)}
        )
        results_workbook_bytes = build_results_workbook(result_sheet_specs)

        auto_scale = auto_scale_factor(coordinates, U)

        st.success("Analisis selesai.")
        col1, col2, col3 = st.columns(3)
        col1.metric("Jumlah node", len(nodes))
        col2.metric("Jumlah elemen", len(elements))
        col3.metric("DOF total", dof)

        col4, col5, col6 = st.columns(3)
        col4.metric("Displacement maksimum", f"{max_disp:.6g}")
        col5.metric("Von Mises maksimum GP (MPa)", f"{max_vm_gauss:.6g}")
        col6.metric(
            "Titik Gauss status Yield",
            f"{yield_point_count}/{len(gauss_results_display_df)}",
        )

        if equilibrium_ok:
            st.success("Cek keseimbangan global gaya: OK.")
        else:
            st.warning(
                "Cek keseimbangan global gaya menunjukkan residual yang perlu ditinjau."
            )

        if yield_point_count > 0:
            st.warning(
                f"Terdapat {yield_point_count} titik Gauss dengan status `Yield` menurut evaluasi Drucker-Prager."
            )
        else:
            st.info("Semua titik Gauss masih berstatus `Elastic` pada evaluasi Drucker-Prager.")

        st.subheader("Ringkasan Eksekutif")
        summary_col1, summary_col2 = st.columns((1.45, 1.0))
        with summary_col1:
            show_dataframe(executive_summary_df.round(6), use_container_width=True)
        with summary_col2:
            st.write("Ringkasan Model")
            show_dataframe(model_summary_df, use_container_width=True)
            st.write("Cek Keseimbangan Gaya")
            show_dataframe(equilibrium_df.round(9), use_container_width=True)
            st.download_button(
                "Unduh workbook hasil analisis",
                data=results_workbook_bytes,
                file_name=RESULTS_FILENAME,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            st.caption(
                "Workbook ini memuat ringkasan, hotspot, input model, hasil utama, vektor sistem, "
                "dan matriks global bila ukurannya masih aman untuk diekspor."
            )

        st.subheader("Hotspot Kritis")
        hotspot_tab1, hotspot_tab2, hotspot_tab3 = st.tabs(
            ["Node & Reaksi", "Elemen", "Titik Gauss"]
        )
        with hotspot_tab1:
            st.write("Node dengan displacement terbesar")
            show_dataframe(nodal_hotspots_df.round(6), use_container_width=True)
            st.write("Tumpuan dengan resultan reaksi terbesar")
            show_dataframe(reaction_hotspots_df.round(6), use_container_width=True)
        with hotspot_tab2:
            st.write("Elemen paling kritis berdasarkan Von Mises")
            show_dataframe(element_vm_hotspots_df.round(6), use_container_width=True)
            st.write("Elemen paling kritis berdasarkan yield function")
            show_dataframe(
                element_yield_hotspots_df.round(6), use_container_width=True
            )
        with hotspot_tab3:
            st.write("Titik Gauss paling kritis berdasarkan Von Mises")
            show_dataframe(gauss_vm_hotspots_df.round(6), use_container_width=True)
            st.write("Titik Gauss paling kritis berdasarkan yield function")
            show_dataframe(gauss_yield_hotspots_df.round(6), use_container_width=True)

        deformation_scale = st.number_input(
            "Skala deformasi plot",
            min_value=0.0,
            value=float(round(auto_scale, 3)),
            format="%.6g",
        )

        figure = plot_mesh(
            nodes,
            elements,
            node_index,
            coordinates,
            U,
            deformation_scale,
            constrained_nodes,
            critical_node_id=critical_node_id,
            critical_element_id=critical_element_id,
        )
        st.caption(
            "Mesh overview menandai node dengan displacement maksimum dan elemen dengan Von Mises elemen tertinggi."
        )
        st.pyplot(figure)

        st.subheader("Animasi Hasil")
        animation_mode = st.selectbox(
            "Jenis animasi",
            [
                "Mesh deformasi",
                "Kontur Uy",
                "Kontur Ux",
                "Kontur tegangan",
            ],
            index=0,
        )
        animation_col1, animation_col2 = st.columns(2)
        animation_frames = animation_col1.slider(
            "Jumlah frame animasi",
            min_value=12,
            max_value=48,
            value=24,
            step=6,
        )
        animation_interval = animation_col2.slider(
            "Kecepatan animasi per frame (ms)",
            min_value=40,
            max_value=240,
            value=90,
            step=10,
        )
        selected_animation_stress_label = None
        if animation_mode == "Kontur tegangan":
            selected_animation_stress_label = st.selectbox(
                "Komponen tegangan animasi",
                list(STRESS_CONTOUR_OPTIONS.keys()),
                index=0,
            )
        st.caption(
            "Animasi ini adalah visualisasi hasil analisis statik yang bergerak dari kondisi awal ke deformasi maksimum lalu kembali lagi. "
            "Ini bukan analisis dinamik atau time history."
        )
        animation_gif = None
        try:
            with st.spinner("Membuat animasi..."):
                if animation_mode == "Mesh deformasi":
                    animation_gif = create_deformation_animation_gif(
                        elements,
                        node_index,
                        coordinates,
                        U,
                        deformation_scale,
                        constrained_nodes,
                        animation_frames,
                        animation_interval,
                    )
                elif animation_mode == "Kontur Uy":
                    animation_gif = create_scalar_contour_animation_gif(
                        elements,
                        node_index,
                        coordinates,
                        U,
                        deformation_scale,
                        constrained_nodes,
                        displacement_values_from_vector(U, "Uy"),
                        "Uy",
                        "satuan model",
                        DISPLACEMENT_CONTOUR_CMAP,
                        animation_frames,
                        animation_interval,
                        "Animasi Kontur Deformasi",
                    )
                elif animation_mode == "Kontur Ux":
                    animation_gif = create_scalar_contour_animation_gif(
                        elements,
                        node_index,
                        coordinates,
                        U,
                        deformation_scale,
                        constrained_nodes,
                        displacement_values_from_vector(U, "Ux"),
                        "Ux",
                        "satuan model",
                        DISPLACEMENT_CONTOUR_CMAP,
                        animation_frames,
                        animation_interval,
                        "Animasi Kontur Deformasi",
                    )
                elif selected_animation_stress_label is not None:
                    animation_stress_column = STRESS_CONTOUR_OPTIONS[
                        selected_animation_stress_label
                    ]
                    animation_gif = create_scalar_contour_animation_gif(
                        elements,
                        node_index,
                        coordinates,
                        U,
                        deformation_scale,
                        constrained_nodes,
                        average_element_result_to_nodes(
                            elements,
                            node_index,
                            element_results_mpa_df,
                            animation_stress_column,
                        ),
                        selected_animation_stress_label,
                        "MPa",
                        STRESS_CONTOUR_CMAP,
                        animation_frames,
                        animation_interval,
                        "Animasi Kontur Tegangan",
                    )
        except Exception as exc:
            st.error(f"Animasi gagal dibuat: {exc}")

        if animation_gif is not None:
            render_gif(animation_gif)

        st.subheader("Kontur Tegangan")
        contour_col1, contour_col2 = st.columns((2, 2))
        selected_stress_label = contour_col1.selectbox(
            "Komponen tegangan",
            list(STRESS_CONTOUR_OPTIONS.keys()),
            index=0,
        )
        contour_geometry = contour_col2.radio(
            "Geometri kontur",
            ["Mesh asli", "Bentuk terdeformasi"],
            index=0,
            horizontal=True,
        )
        stress_figure = plot_stress_contour(
            elements,
            node_index,
            coordinates,
            U,
            deformation_scale,
            constrained_nodes,
            element_results_mpa_df,
            STRESS_CONTOUR_OPTIONS[selected_stress_label],
            contour_geometry == "Bentuk terdeformasi",
        )
        if stress_figure is not None:
            st.caption(
                "Gradasi warna memakai kuning untuk tegangan terendah dan merah untuk tegangan tertinggi. "
                "Nilai kontur diperoleh dari hasil tegangan elemen yang diinterpolasi ke node."
            )
            st.pyplot(stress_figure)

        st.subheader("Kontur Deformasi")
        displacement_contour_geometry = st.radio(
            "Geometri kontur deformasi",
            ["Mesh asli", "Bentuk terdeformasi"],
            index=1,
            horizontal=True,
        )
        uy_figure = plot_displacement_contour(
            elements,
            node_index,
            coordinates,
            U,
            deformation_scale,
            constrained_nodes,
            DISPLACEMENT_CONTOUR_OPTIONS["Uy"],
            displacement_contour_geometry == "Bentuk terdeformasi",
        )
        ux_figure = plot_displacement_contour(
            elements,
            node_index,
            coordinates,
            U,
            deformation_scale,
            constrained_nodes,
            DISPLACEMENT_CONTOUR_OPTIONS["Ux"],
            displacement_contour_geometry == "Bentuk terdeformasi",
        )
        st.caption(
            "Kontur deformasi ditampilkan satu per satu agar area gambar lebih besar. "
            "Nilai terendah berada pada gradasi biru dan nilai maksimum ditampilkan merah."
        )
        if uy_figure is not None:
            st.pyplot(uy_figure)
        if ux_figure is not None:
            st.pyplot(ux_figure)

        st.subheader("Displacement Nodal")
        show_dataframe(displacement_df.round(6), use_container_width=True)

        st.subheader("Reaksi Tumpuan")
        show_dataframe(
            constrained_reaction_display_df.round(6), use_container_width=True
        )

        st.subheader("Output Elemen Lengkap")
        st.caption(
            "Hasil elemen dihitung langsung dengan `B_total = [B]1 + [B]2 + [B]3 + [B]4` "
            "agar konsisten dengan workbook Excel. Kolom strain tetap tak berdimensi, "
            "sedangkan semua kolom tegangan pada tabel ini ditampilkan dalam MPa. "
            "Kolom `principal_angle_*_ccw_deg` memakai konvensi positif berlawanan arah jarum jam. "
            "Kolom `principal_angle_plot_direction` mengikuti aturan: sudut `+` = berlawanan arah jarum jam, "
            "sudut `-` = searah jarum jam. Kolom `principal_angle_perpendicular_deg` adalah arah tegak lurus terhadap `sigma1`."
        )
        show_dataframe(element_results_display_df.round(6), use_container_width=True)

        principal_sketch_figure = plot_principal_stress_sketches(
            element_results_mpa_df
        )
        if principal_sketch_figure is not None:
            st.subheader("Sketsa Tegangan Utama per Elemen")
            st.caption(
                "Sketsa ini memakai nilai `principal_angle_1_deg` langsung sebagai arah `sigma1`. "
                "Jika nilainya positif, arah digambar berlawanan arah jarum jam (CCW) dari sumbu `+x`, "
                "dan jika negatif digambar searah jarum jam. Jika `sigma1 > 0`, sudut retak dihitung dari arah tegak lurus `sigma1`; "
                "jika `sigma1 < 0`, ditampilkan bahwa retak tarik primer tidak terjadi."
            )
            st.pyplot(principal_sketch_figure)

        mohr_figure = plot_mohr_circles(element_results_mpa_df)
        if mohr_figure is not None:
            st.subheader("Lingkaran Mohr per Elemen")
            st.caption(
                "Plot ini memakai tegangan elemen hasil `B_total` (`sx`, `sy`, `txy`) "
                "dalam MPa, "
                "sehingga setiap elemen direpresentasikan oleh satu lingkaran Mohr."
            )
            st.pyplot(mohr_figure)

        with st.expander("Lingkaran Mohr per Titik Gauss", expanded=False):
            st.caption(
                "Setiap elemen ditampilkan dengan 4 lingkaran Mohr sesuai 4 titik Gauss. "
                "Ini lebih detail dibanding hasil elemen berbasis `B_total`."
            )
            for element_id in sorted(gauss_results_mpa_df["element"].unique()):
                with st.expander(f"Elemen {int(element_id)}", expanded=False):
                    gauss_mohr_figure = plot_mohr_circles_for_element_gauss(
                        gauss_results_mpa_df, int(element_id)
                    )
                    if gauss_mohr_figure is not None:
                        st.pyplot(gauss_mohr_figure)

        with st.expander("Sketsa Tegangan Utama per Titik Gauss", expanded=False):
            st.caption(
                "Sketsa titik Gauss juga memakai `principal_angle_1_deg` langsung untuk arah `sigma1`, "
                "dan arah tegak lurus dibentuk otomatis 90 derajat terhadap `sigma1`. "
                "Aturan retak tarik primer mengikuti tanda `sigma1`."
            )
            for element_id in sorted(gauss_results_mpa_df["element"].unique()):
                with st.expander(f"Elemen {int(element_id)}", expanded=False):
                    gauss_principal_figure = (
                        plot_principal_stress_sketches_for_element_gauss(
                            gauss_results_mpa_df, int(element_id)
                        )
                    )
                    if gauss_principal_figure is not None:
                        st.pyplot(gauss_principal_figure)

        with st.expander("Hasil Regangan dan Tegangan di Titik Gauss", expanded=False):
            st.caption(
                "Tabel ini memuat strain, stress, principal stress/angle, shear extreme, "
                "Von Mises, Drucker-Prager, dan fungsi luluh pada setiap titik Gauss. "
                "Semua kolom tegangan pada tabel ini sudah dalam MPa. "
                "Kolom `principal_angle_*_ccw_deg` memakai konvensi positif berlawanan arah jarum jam. "
                "Kolom `principal_angle_plot_direction` mengikuti aturan arah plot `sigma1`, "
                "dan `principal_angle_perpendicular_deg` adalah arah tegak lurus terhadap `sigma1`."
            )
            show_dataframe(gauss_results_display_df.round(6), use_container_width=True)

        with st.expander("Gaya Nodal Total", expanded=False):
            total_force_df = pd.DataFrame(
                [{"Sum_Fx": F[0::2].sum(), "Sum_Fy": F[1::2].sum()}]
            )
            show_dataframe(total_force_df.round(6), use_container_width=True)

        with st.expander("Matriks Global dan Vektor Sistem", expanded=False):
            st.write("Matrik Beban Nodal `{P}`")
            show_dataframe(
                vector_to_dataframe(F, global_labels, "P").round(6),
                use_container_width=True,
            )

            st.write("Matrik Nodal Displacement `{d}`")
            show_dataframe(
                vector_to_dataframe(U, global_labels, "d").round(6),
                use_container_width=True,
            )

            st.write("Matrik Reaksi Perletakan `{R}`")
            show_dataframe(
                vector_to_dataframe(reactions, global_labels, "R").round(6),
                use_container_width=True,
            )

            st.write("Matrik Kekakuan Global Total `[KG]`")
            show_dataframe(
                matrix_to_dataframe(K, global_labels, global_labels).round(6),
                keep_left_column=True,
                use_container_width=True,
                height=500,
            )

        with st.expander("Matriks Detail per Elemen", expanded=False):
            st.caption(
                "`Matrik Total Koefisien [B]` ditampilkan sebagai penjumlahan "
                "`[B]1 + [B]2 + [B]3 + [B]4` agar sama dengan workbook Excel. "
                "`Matrik Konstitutif [D]` ditampilkan dalam MPa, sedangkan matriks kekakuan "
                "tetap mengikuti satuan sistem input."
            )
            for element_data in element_debug_data:
                element_id = element_data["element_id"]
                local_labels = element_data["local_labels"]
                node_list = ", ".join(str(node_id) for node_id in element_data["node_ids"])

                with st.expander(
                    f"Elemen {element_id} | node [{node_list}]", expanded=False
                ):
                    st.write("Matrik Konstitutif `[D]` (MPa)")
                    show_dataframe(
                        matrix_to_dataframe(
                            convert_stress_matrix_to_mpa(element_data["D"]),
                            ["sx", "sy", "txy"],
                            ["ex", "ey", "gxy"],
                        ).round(6),
                        keep_left_column=True,
                        use_container_width=True,
                    )

                    for gauss_data in element_data["gauss_data"]:
                        gauss_index = gauss_data["gauss_point"]
                        st.write(
                            f"Titik Gauss / node integrasi {gauss_index} "
                            f"(xi={gauss_data['xi']:.4f}, eta={gauss_data['eta']:.4f})"
                        )
                        st.write("Matrik Koefisien `[B]`")
                        show_dataframe(
                            matrix_to_dataframe(
                                gauss_data["B"],
                                ["ex", "ey", "gxy"],
                                local_labels,
                            ).round(6),
                            keep_left_column=True,
                            use_container_width=True,
                        )
                        st.write("Hasil `[B]^T[D][B]`")
                        show_dataframe(
                            matrix_to_dataframe(
                                gauss_data["BTDB"],
                                local_labels,
                                local_labels,
                            ).round(6),
                            keep_left_column=True,
                            use_container_width=True,
                        )

                    st.write("Matrik Total Koefisien `[B]`")
                    show_dataframe(
                        matrix_to_dataframe(
                            element_data["B_total"],
                            b_total_row_labels(),
                            local_labels,
                        ).round(6),
                        keep_left_column=True,
                        use_container_width=True,
                    )

                    st.write("Matrik Kekakuan Lokal Elemen `[ke]`")
                    show_dataframe(
                        matrix_to_dataframe(
                            element_data["ke"],
                            local_labels,
                            local_labels,
                        ).round(6),
                        keep_left_column=True,
                        use_container_width=True,
                    )

                    st.write("Matrik Kekakuan Global Elemen `[KG]`")
                    show_dataframe(
                        matrix_to_dataframe(
                            element_data["KG_element"],
                            global_labels,
                            global_labels,
                        ).round(6),
                        keep_left_column=True,
                        use_container_width=True,
                        height=400,
                    )

    except Exception as exc:
        st.error(str(exc))
